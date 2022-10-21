# Copyright (c) 2022, Teampro and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)


class ItemCurrencyConversion(Document):
    @frappe.whitelist()
    def get_data(self):
        data = """<table class='table table-bordered=1'>
		<tr>
		   <td style="background-color:#e22026; padding:2px; color:white;border: 1px solid black; font-size:15px;">
				 <center><b>S.No</b></center>
		   <td style="background-color:#e22026; padding:2px;color:white; border: 1px solid black; font-size:15px;">
				 <center><b>Item Code</b></center>
		   <td style="background-color:#e22026; padding:2px; color:white;border: 1px solid black; font-size:15px;">
				 <center><b>Item Name</b></center>
		   <td style="background-color:#e22026; padding:2px;color:white; border: 1px solid black; font-size:15px;">
				 <center><b>Price List</b></center>
		   <td style="background-color:#e22026;color:white; padding:2px; border: 1px solid black; font-size:15px;">
				 <center><b>Default Currency</b></center>
		   <td style="background-color:#e22026; padding:2px; color:white;border: 1px solid black; font-size:15px;">
				 <center><b>Default rate</b></center>
		   <td style="background-color:#e22026; padding:2px; color:white;border: 1px solid black; font-size:15px;">
				 <center><b>Converted Currency</b></center>
		   <td style="background-color:#e22026; padding:2px;color:white; border: 1px solid black; font-size:15px;">
				 <center><b>Converted Rate</b></center>
		</tr>"""
        # rate = frappe.db.get_value('Currency Exchange', {
        #                            "from_currency": self.from_currency, "to_currency": self.convert_to}, ['exchange_rate'])
        # print(rate)

        price = frappe.get_all(
            'Item Price', {"price_list": self.price_list}, ['*'])
        rate = frappe.db.get_value('Currency Exchange', {
                                   "from_currency": self.from_currency, "to_currency": self.convert_to}, ['exchange_rate'])
        frappe.errprint(rate)
        # price = frappe.get_all('Item Price',  ['*'])
        i = 1
        for p in price:
            data += """
			  <tr>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
					<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>


			  </tr>
			""" % (i, p.item_code, p.item_name, p.price_list, p.currency, p.price_list_rate,self.convert_to,p.price_list_rate * rate)
            i += 1
        return data


@frappe.whitelist()
def download():
    filename = 'Item Currency Conversion'
    test = build_xlsx_response(filename)


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    # frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["s.no", "Item Code", "Item Name", "Price List",
              "Default Currency", "Default Rate", "Converted Currency", "Converted Rate"]
    ws.append(header)
    # rate = frappe.db.get_value('Currency Exchange', {
    #                            "from_currency": args.from_currency, "to_currency": args.convert_to}, 'exchange_rate')
    data = []

    price = frappe.get_all(

        'Item Price', {"price_list": args.price_list}, ['*'])
    rate = frappe.db.get_value('Currency Exchange', {
                               "from_currency": args.from_currency, "to_currency": args.convert_to}, 'exchange_rate')

    i = 1
    for p in price:
        data.append([i, p.item_code, p.item_name, p.price_list,
                    p.currency, p.price_list_rate, args.convert_to, (rate * p.price_list_rate)])
        # frappe.log_errror(p.item_code)

    i += 1
    for p in data:

        ws.append(p)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=18):
    #     for cell in rows:
    #         cell.fill = PatternFill( fill_type = "solid")

    bold_font = Font(bold=True, size=12)
    for cell in ws["1:1"]:
        cell.font = bold_font

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


# def get_price(args):
#     price = frappe.get_all(
#         'Item Price', {"price_list": args.price_list}, ['*'])
#     data = []
#     i = 1
#     for p in price:
#         data.append([i, p.item_code, p.item_name, p.price_list,
#                     p.currency, p.price_list_rate, "demo", "demo"])
#         frappe.errprint(p.item_code)

#     i += 1

#     return data
