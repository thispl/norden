# Copyright (c) 2022, Abdulla and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)

class ProductSearch(Document):
	@frappe.whitelist()
	def get_data(self):
		data = ''
		data1 = ''
		item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
		rate = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
		group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
		des = frappe.get_value('Item',{'item_code':self.item_code},'description')
		price = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Cost'},'price_list_rate')
		c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
		csp = 'Current Selling Price'
		cpp = 'Current Purchase Price'
		cost = 'COST'
		pso = 'Pending Sales order'
		po ='Total Purchase Order'
		ppo = 'Pending Purchase order'
		cspp_rate = 0
		cppp_rate = 0
		psoc = 0
		ppoc = 0
		ppoc_total = 0
		i = 0
		cou = 0
		p_po = 0
		p_so = 0
		tot = 'Total'
		uom = 'Nos'

		stocks_query = frappe.db.sql("""select actual_qty,reserved_stock,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' """%(item),as_dict=True)
		# frappe.errprint(stocks_query.reserved_stock)
		if stocks_query:
			stocks = stocks_query
		
		data += '<table class="table table-bordered" style="width:70%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white" colspan=10><center>NORDEN PRODUCT SEARCH</center></th></tr>'
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
		data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(des)
		
		if stocks_query:
			data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Cost</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Selling Rate</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Currency</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending PO</b></center></td>'
			data += '<td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending SO</b></center></td>'
			data += '</tr>'
			for stock in stocks_query:
				if stock.actual_qty >= 0:
					reserve = stock.actual_qty - stock.reserved_stock
					stock_company = frappe.db.sql("""select company from tabWarehouse where name = '%s' """%(stock.warehouse),as_dict=True)
					for com in stock_company:
						psoc_query = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty from `tabSales Order`
						left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
						where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' """ % (self.item_code,com.company), as_dict=True)[0]
						if not psoc_query["qty"]:
							psoc_query["qty"] = 0
						deliver = frappe.db.sql("""select sum(`tabDelivery Note Item`.qty) as qty from `tabDelivery Note`
						left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
						where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.company = '%s'  """%(self.item_code,com.company), as_dict=True)[0]
						if not deliver["qty"]:
							deliver['qty'] = 0
						del_total = psoc_query['qty'] - deliver['qty']
						ppoc_query = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
						left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
						where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 and `tabPurchase Order`.company = '%s' """ % (self.item_code,com.company), as_dict=True)[0]
						if not ppoc_query["qty"]:
							ppoc_query["qty"] = 0
						ppoc_receipt = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.qty) as qty from `tabPurchase Receipt`
						left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
						where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt`.status = "Completed" and `tabPurchase Receipt`.company = '%s'  """%(self.item_code,com.company),as_dict=True)[0]
						if not ppoc_receipt["qty"]:
							ppoc_receipt["qty"] = 0
						ppoc_total = ppoc_query["qty"] - ppoc_receipt["qty"]
						country,default_currency = frappe.get_value("Company",{"name":com.company},["country","default_currency"])
						if country == "United Arab Emirates":
							cost = frappe.get_value("Item Price",{"item_code":item,"price_list":"Electra Qatar - NCMEF"},["price_list_rate"])
						else:
							cost = frappe.get_value("Item Price",{"item_code":item,"price_list":"STANDARD BUYING-USD"},["price_list_rate"])
						
						valuation_rate = 0
						source_warehouse = frappe.db.get_value('Warehouse', {'default_for_stock_transfer':1,'company': com.company }, ["name"])
						latest_vr = frappe.db.sql("""select valuation_rate as vr from tabBin
								where item_code = '%s' and warehouse = '%s' """%(item,source_warehouse),as_dict=True)
						if latest_vr:
							valuation_rate = latest_vr[0]["vr"]
						else:
							val_rate = []
							l_vr = frappe.db.sql("""
								SELECT valuation_rate AS vr FROM tabBin
								WHERE item_code = %s AND warehouse = %s
							""", (item, source_warehouse), as_dict=True)
							for item in l_vr: 
								if item not in val_rate: 
									val_rate.append(item.vr)
							if len(val_rate) > 1 :
								valuation_rate = max(val_rate)
					     
					     
						pricelist = country + ' ' + "Sales Price"
						if country == "United Arab Emirates":
							sp = frappe.get_value("Item Price",{"item_code":item,"price_list":"Internal - NCMEF"},["price_list_rate"])
						else:
							sp = frappe.get_value("Item Price",{"item_code":item,"price_list":pricelist},["price_list_rate"])
						data += '<tr><td colspan = 1 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black">%s</td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(com.company,stock.warehouse,reserve or 0,stock.stock_uom or '-',valuation_rate or 0,sp or 0,default_currency,ppoc_total or 0,del_total or 0)
						i += 1
						# cou += stock.actual_qty
						cou += reserve
						p_po += ppoc_total
						p_so += del_total
			data += '<tr><td align="right" colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 3 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b></b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 1 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td></tr>'%(tot or 0,int(cou) or 0,uom,p_po or 0,p_so or 0)
			data += '</table>'
		else:
			i += 1
			data1 += '<tr><td align="center" colspan = 10 style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white;"><b>No Stock Available</b></td></tr>'
			data1 += '</table>'
			data += data1
		if i > 0:
			return data
		
	
	@frappe.whitelist()
	def get_data_norden(self):
		data = ''
		data1 = ''
		aa = self.item_code
		item = frappe.get_value('Item',{'item_code':self.item_code},'item_code')
		rate = frappe.get_value('Item',{'item_code':self.item_code},'valuation_rate')
		group = frappe.get_value('Item',{'item_code':self.item_code},'item_group')
		des = frappe.get_value('Item',{'item_code':self.item_code},'description')
		price = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Cost'},'price_list_rate')
		c_s_p = frappe.get_value('Item Price',{'item_code':self.item_code,'price_list':'Standard Selling'},'price_list_rate') or 0
		csp = 'Current Selling Price'
		cpp = 'Current Purchase Price'
		cost = 'COST'
		pso = 'Pending Sales order'
		po ='Total Purchase Order'
		ppo = 'Pending Purchase order'
		cspp_rate = 0
		cppp_rate = 0
		psoc = 0
		ppoc = 0
		ppoc_total = 0
		i = 0
		cou = 0
		p_po = 0
		p_so = 0
		tot = 'Total'
		uom = 'Nos'

		stocks_query = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
				where item_code = '%s' """%(item),as_dict=True)
		if stocks_query:
			stocks = stocks_query
		if ppoc_total > 0 :
			ppoc_date_query = frappe.db.sql("""select `tabPurchase Order Item`.schedule_date  as date from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' """ % (self.item_code), as_dict=True)[0]
			if ppoc_date_query:
				po_date = ppoc_date_query['date']

		data += '<table class="table table-bordered" style="width:70%"><tr><th style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white" colspan=12><center>NORDEN - PRODUCT SEARCH</center></th></tr>'
		data += '<tr><td colspan = 4 style="padding:1px;border: 1px solid black;color:white;background-color:#6f6f6f;text-align: left"><b>Item Code</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(item)
		data += '<tr><td colspan = 4 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Name</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(frappe.db.get_value('Item',item,'item_name'))
		data += '<tr><td colspan = 4 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Item Group</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(group)
		data += '<tr><td colspan = 4 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;text-align: left"><b>Description</b></td><td colspan = 8 style="padding:1px;border: 1px solid black;text-align: left"><b>%s</b></td></tr>'%(des)
		if stocks_query:
			data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Company</b></center></td>'
			data += '<td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Warehouse</b></center></td>'
			data += '<td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>QTY</b></center></td>'
			data += '<td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>UOM</b></center></td>'
			data += '<td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending PO</b></center></td>'
			data += '<td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>Pending SO</b></center></td>'
			data += '</tr>'
			for stock in stocks_query:
				if stock.actual_qty >= 0:
					stock_company = frappe.db.sql("""select company from tabWarehouse where name = '%s' """%(stock.warehouse),as_dict=True)
					for com in stock_company:
						psoc_query = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty from `tabSales Order`
						left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
						where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 1 and `tabSales Order`.company = '%s' """ % (self.item_code,com.company), as_dict=True)[0]
						if not psoc_query["qty"]:
							psoc_query["qty"] = 0
						deliver = frappe.db.sql("""select sum(`tabDelivery Note Item`.qty) as qty from `tabDelivery Note`
						left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
						where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.company = '%s'  """%(self.item_code,com.company), as_dict=True)[0]
						if not deliver["qty"]:
							deliver['qty'] = 0
						del_total = psoc_query['qty'] - deliver['qty']
						ppoc_query = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
						left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
						where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 and `tabPurchase Order`.company = '%s' """ % (self.item_code,com.company), as_dict=True)[0]
						if not ppoc_query["qty"]:
							ppoc_query["qty"] = 0
						ppoc_receipt = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.qty) as qty from `tabPurchase Receipt`
						left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
						where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt`.status = "Completed" and `tabPurchase Receipt`.company = '%s'  """%(self.item_code,com.company),as_dict=True)[0]
						if not ppoc_receipt["qty"]:
							ppoc_receipt["qty"] = 0
						ppoc_total = ppoc_query["qty"] - ppoc_receipt["qty"]
						country,default_currency = frappe.get_value("Company",{"name":com.company},["country","default_currency"])
						pricelist = country + ' ' + "Sales Price"
						if country == "United Arab Emirates":
							sp = frappe.get_value("Item Price",{"item_code":item,"price_list":"Internal - NCMEF"},["price_list_rate"])
						else:
							sp = frappe.get_value("Item Price",{"item_code":item,"price_list":pricelist},["price_list_rate"])
						if stock.actual_qty == 0 and sp == 0 and ppoc_total == 0 and del_total == 0:
							frappe.errprint(stock.actual_qty)
						else:
							data += '<tr><td colspan = 2 style="padding:1px;border: 1px solid black">%s</td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black"><center><b>%s</b></center></td></tr>'%(com.company,stock.warehouse,int(stock.actual_qty) or 0,stock.stock_uom or '-',ppoc_total or 0,del_total or 0)
						i += 1
						cou += stock.actual_qty
						p_po += ppoc_total
						p_so += del_total
			data += '<tr><td align="right" colspan = 4 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td align="right" colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><b>%s</b></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td><td colspan = 2 style="padding:1px;border: 1px solid black;background-color:#6f6f6f;color:white;"><center><b>%s</b></center></td></tr>'%(tot or 0,int(cou) or 0,uom,p_po or 0,p_so or 0)
			data += '</table>'
		else:
			i += 1
			data1 += '<tr><td align="center" colspan = 10 style="padding:1px;border: 1px solid black;background-color:#fe3f0c;color:white;"><b>No Stock Available</b></td></tr>'
			data1 += '</table>'
			data += data1
		if i > 0:
			return data
