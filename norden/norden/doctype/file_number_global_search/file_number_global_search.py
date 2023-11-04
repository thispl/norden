# Copyright (c) 2022, Teampro and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
	flt,
	cint,
	cstr,
	get_html_format,
	get_url_to_form,
	gzip_decompress,
	format_duration,
)


class FileNumberGlobalSearch(Document):
	@frappe.whitelist()
	def get_data(self):
		if self.file_number:
			quotations = frappe.get_all('Quotation', filters={'file_number': self.file_number}, fields=['name', 'party_name', 'file_number', 'grand_total', 'status', 'transaction_date'])
			data = ''
			data += '<table class="table table-bordered" width="100%">'
			data += '<tr rowspan="2"><th style="padding:1px; border: 1px solid black;" colspan="18" width="100%"><center>Linked Documents</center></th></tr>'
			data += '<tr colspan="1" style =font-size:10px><th style="border: 1px solid black;" colspan="2" width="12%"><center>File Number</center></th><th style="border: 1px solid black;" colspan="2" width="11%"><center>Party</center></th><th style="border: 1px solid black;" colspan="2" width="11%"><center>Quotation</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Sales Order</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Delivery Note</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Sales Invoice</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Material Request</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Purchase Order</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Purchase Receipt</center></th></tr>'
			td_style = 'style="border: 1px solid black; font-size:9px; padding: 5px;"'
			for quotation in quotations:
				if quotation['status'] != "Cancelled":
					data += '<tr style="font-size:9px"><td style="border: 1px solid black;" colspan="2"><center>%s</center></td>' % quotation['file_number']
					data += '<td style="border: 1px solid black;" colspan="2"><center>%s</center></td>' % quotation['party_name']

					if quotation['name']:
						data += '<td style="border: 1px solid black;" colspan="2">' + \
								'<center><a href="https://erp.nordencommunication.com/app/quotation/%s"><b>%s</b></a></center>' % (quotation['name'], quotation['name']) + \
								'<center>%s</center>' % quotation['grand_total'] + \
								'<center>%s</center>' % quotation['status'] + \
								'<center>%s</center>' % quotation['transaction_date'] + \
								'</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					sales_orders = frappe.get_all('Sales Order', { 'file_number': self.file_number}, ['name', 'grand_total', 'status', 'transaction_date'])
					so_values = []
					for so in sales_orders:
						if so['status'] != "Cancelled":
							so_values.append('<center><a href="https://erp.nordencommunication.com/app/sales-order/%s"><b>%s</b></a></center>' % (so['name'], so['name']) + \
											'<center>%s</center>' % so['grand_total'] + \
											'<center>%s</center>' % so['status'] + \
											'<center>%s</center>' % so['transaction_date'])
					if so_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(so_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					# Add similar sections for Delivery Note
					delivery_notes = frappe.get_all('Delivery Note', { 'file_number': self.file_number}, ['name', 'grand_total', 'status', 'posting_date'])
					dn_values = []
					for dn in delivery_notes:
						if dn['status'] != 'Cancelled':
							dn_values.append('<center><a href="https://erp.nordencommunication.com/app/delivery-note/%s"><b>%s</b></a></center>' % (dn['name'], dn['name']) + \
											'<center>%s</center>' % dn['grand_total'] + \
											'<center>%s</center>' % dn['status'] + \
											'<center>%s</center>' % dn['posting_date'])
					if dn_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(dn_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					# Add similar sections for Sales Invoice, Material Request, Purchase Order, and Purchase Receipt
					sales_invoices = frappe.get_all('Sales Invoice', { 'file_number': self.file_number}, ['name', 'grand_total', 'status', 'posting_date'])
					si_values = []
					for si in sales_invoices:
						if si['status'] != "Cancelled":
							si_values.append('<center><a href="https://erp.nordencommunication.com/app/sales-invoice/%s"><b>%s</b></a></center>' % (si['name'], si['name']) + \
											'<center>%s</center>' % si['grand_total'] + \
											'<center>%s</center>' % si['status'] + \
											'<center>%s</center>' % si['posting_date'])
					if si_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(si_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					material_requests = frappe.get_all('Material Request', { 'file_number': self.file_number}, ['name', 'total', 'status', 'transaction_date'])
					mr_values = []
					for mr in material_requests:
						if mr['status'] != "Cancelled":
							mr_values.append('<center><a href="https://erp.nordencommunication.com/app/material-request/%s"><b>%s</b></a></center>' % (mr['name'], mr['name']) + \
											'<center>%s</center>' % mr['total'] + \
											'<center>%s</center>' % mr['status'] + \
											'<center>%s</center>' % mr['transaction_date'])
					if mr_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(mr_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					purchase_orders = frappe.get_all('Purchase Order', { 'file_number': self.file_number}, ['name', 'grand_total', 'status', 'transaction_date'])
					po_values = []
					for po in purchase_orders:
						if po['status'] != "Cancelled":
							po_values.append('<center><a href="https://erp.nordencommunication.com/app/purchase-order/%s"><b>%s</b></a></center>' % (po['name'], po['name']) + \
											'<center>%s</center>' % po['grand_total'] + \
											'<center>%s</center>' % po['status'] + \
											'<center>%s</center>' % po['transaction_date'])
					if po_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(po_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
					purchase_receipts = frappe.get_all('Purchase Receipt', { 'file_number': self.file_number}, ['name', 'grand_total', 'status', 'posting_date'])
					pr_values = []
					for pr in purchase_receipts:
						if pr['status'] != "Cancelled":
							pr_values.append('<center><a href="https://erp.nordencommunication.com/app/purchase-receipt/%s"><b>%s</b></a></center>' % (pr['name'], pr['name']) + \
											'<center>%s</center>' % pr['grand_total'] + \
											'<center>%s</center>' % pr['status'] + \
											'<center>%s</center>' % pr['posting_date'])
					if pr_values:
						data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(pr_values) + '</td>'
					else:
						data += '<td style="border: 1px solid black;" colspan="2"></td>'
					
				data += '</tr>'
			data += '</table>'
			return data

	
	@frappe.whitelist()
	def get_data_value(self):
		quotations = frappe.get_all('Quotation', filters={'company': self.company}, fields=['name', 'party_name', 'file_number', 'grand_total', 'status', 'transaction_date'])
		data = ''
		data += '<table class="table table-bordered" width="100%">'
		data += '<tr rowspan="2"><th style="padding:1px; border: 1px solid black;" colspan="18" width="100%"><center>Linked Documents</center></th></tr>'
		data += '<tr style =font-size:10px><th style="border: 1px solid black;" colspan="2" width="12%"><center>File Number</center></th><th style="border: 1px solid black;" colspan="2" width="11%"><center>Party</center></th><th style="border: 1px solid black;" colspan="2" width="11%"><center>Quotation</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Sales Order</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Delivery Note</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Sales Invoice</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Material Request</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Purchase Order</center><th style="border: 1px solid black;" colspan="2" width="11%"><center>Purchase Receipt</center></th></tr>'
		td_style = 'style="border: 1px solid black; font-size:9px; padding: 5px;"'
		for quotation in quotations:
			if quotation['status'] != "Cancelled":
				data += '<tr style="font-size:9px"><td style="border: 1px solid black;" colspan="2"><center>%s</center></td>' % quotation['file_number']
				data += '<td style="border: 1px solid black;" colspan="2"><center>%s</center></td>' % quotation['party_name']

				if quotation['name']:
					data += '<td style="border: 1px solid black;" colspan="2">' + \
							'<center><a href="https://erp.nordencommunication.com/app/quotation/%s"><b>%s</b></a></center>' % (quotation['name'], quotation['name']) + \
							'<center>%s</center>' % quotation['grand_total'] + \
							'<center>%s</center>' % quotation['status'] + \
							'<center>%s</center>' % quotation['transaction_date'] + \
							'</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

				sales_order = frappe.get_all('Sales Order', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'grand_total', 'status', 'transaction_date'])
				so_values = []
				for so in sales_order:
					if so['status'] != "Cancelled":
						so_values.append('<center><a href="https://erp.nordencommunication.com/app/sales-order/%s"><b>%s</b></a></center>' % (so['name'], so['name']) + \
										'<center>%s</center>' % so['grand_total'] + \
										'<center>%s</center>' % so['status'] + \
										'<center>%s</center>' % so['transaction_date'])
				if so_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(so_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

			   

				delivery = frappe.get_all('Delivery Note', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'grand_total', 'status', 'posting_date'])
				dn_values = []
				for dn in delivery:
					if dn['status'] != 'Cancelled':
						dn_values.append('<center><a href="https://erp.nordencommunication.com/app/delivery-note/%s"><b>%s</b></a></center>' % (dn['name'], dn['name']) + \
										'<center>%s</center>' % dn['grand_total'] + \
										'<center>%s</center>' % dn['status'] + \
										'<center>%s</center>' % dn['posting_date'])
				if dn_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(dn_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

				sales_invoice = frappe.get_all('Sales Invoice', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'grand_total', 'status', 'posting_date'])
				si_values = []
				for si in sales_invoice:
					if si['status'] != "Cancelled":
						si_values.append('<center><a href="https://erp.nordencommunication.com/app/sales-invoice/%s"><b>%s</b></a></center>' % (si['name'], si['name']) + \
										'<center>%s</center>' % si['grand_total'] + \
										'<center>%s</center>' % si['status'] + \
										'<center>%s</center>' % si['posting_date'])
				if si_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(si_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

				material_request = frappe.get_all('Material Request', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'total', 'status', 'transaction_date'])
				mr_values = []
				for mr in material_request:
					if mr['status'] != "Cancelled":
						mr_values.append('<center><a href="https://erp.nordencommunication.com/app/material-request/%s"><b>%s</b></a></center>' % (mr['name'], mr['name']) + \
										'<center>%s</center>' % mr['total'] + \
										'<center>%s</center>' % mr['status'] + \
										'<center>%s</center>' % mr['transaction_date'])
				if si_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(mr_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'
				
				purchase_order = frappe.get_all('Purchase Order', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'grand_total', 'status', 'transaction_date'])
				po_values = []
				for po in purchase_order:
					if po['status'] != "Cancelled":
						po_values.append('<center><a href="https://erp.nordencommunication.com/app/purchase-order/%s"><b>%s</b></a></center>' % (po['name'], po['name']) + \
										'<center>%s</center>' % po['grand_total'] + \
										'<center>%s</center>' % po['status'] + \
										'<center>%s</center>' % po['transaction_date'])
				if po_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(po_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

				receipt = frappe.get_all('Purchase Receipt', {'company': self.company, 'file_number': quotation['file_number']}, ['name', 'grand_total', 'status', 'posting_date'])
				receipt_values = []
				for pay in receipt:
					if pay['status'] != "Cancelled":
						receipt_values.append('<center><a href="https://erp.nordencommunication.com/app/purchase-receipt/%s"><b>%s</b></a></center>' % (pay['name'], pay['name']) + \
										'<center>%s</center>' % pay['grand_total'] + \
										'<center>%s</center>' % pay['status'] + \
										'<center>%s</center>' % pay['posting_date'])
				if receipt_values:
					data += '<td style="border: 1px solid black;font-size:9px" colspan="2">' + ' '.join(receipt_values) + '</td>'
				else:
					data += '<td style="border: 1px solid black;" colspan="2"></td>'

				data += '</tr>'

		data += '</table>'
		return data
