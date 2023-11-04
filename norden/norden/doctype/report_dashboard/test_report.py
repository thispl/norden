import frappe
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)


@frappe.whitelist()
def download():
    filename = 'Payroll Summary Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    att_date = getdate(args.from_date)
    att_month = att_date.strftime("%B")
    att_year = att_date.year

    ws = wb.create_sheet(sheet_name, 0)  

    ws.append([str(att_month)+'-'+str(att_year)])
    
    ws.append(['Additions','','','','','Deductions','','','','','',''])
    
    ws.append(['Particulars','Number of Staffs','Gross','Arrears/Other Additions','Birthday Bonus','Mobile & Internet Allowance','Professional Tax','EPF Deduction','LWF','TDS-Salary','Salary Advance','Net Amount'])
    total_staff_count = 0
    total_arrear_amount = 0
    total_gross_amount = 0
    total_birthday_amount = 0
    total_mobile_amount = 0
    total_pf_ammount = 0
    total_epf_amount = 0
    total_lwf_amount = 0
    total_tds_amount = 0
    total_salary_advance = 0
    total_net_pay = 0
    category = frappe.db.sql("""select * from `tabPayroll Category` """,as_dict=1)
    for c in category:
        staff_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` 
        ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = '%s' 
        AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(c.name,args.from_date,args.to_date),as_dict=True)[0] or 0.0
        if staff_count['count'] != None:
            total_staff_count += staff_count['count']

        arrear_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` 
        INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
        INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent 
        WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = '%s' 
        AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' 
        AND `tabSalary Detail`.salary_component = 'Arrear'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if arrear_amount['amount'] != None:
            total_arrear_amount += arrear_amount['amount']

        gross_amount = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS amount FROM `tabSalary Slip` 
        INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
        WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = '%s'
        AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if gross_amount['amount'] != None:
            total_gross_amount += gross_amount['amount']

        birthday_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount 
        FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name
        INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
        WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = '%s' 
        AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'
        AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if birthday_amount['amount'] != None:
            total_birthday_amount += birthday_amount['amount']

        mobile_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` 
        INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN 
        `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent 
        WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = '%s' 
        AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'
        AND `tabSalary Detail`.salary_component = 'Internet'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if mobile_amount['amount'] != None:
            total_mobile_amount += mobile_amount['amount']

        pf_ammount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` 
        INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
        INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE `tabEmployee`.status = 'Active' 
        AND `tabEmployee`.payroll_category = '%s' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' 
        AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if pf_ammount['amount'] != None:
            total_pf_ammount += pf_ammount['amount']

        epf_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` INNER JOIN `tabEmployee`
        ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON 
        `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active'
        AND `tabEmployee`.payroll_category = '%s' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'
        AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if epf_amount['amount'] != None:
            total_epf_amount += epf_amount['amount']

        lwf_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` INNER JOIN `tabEmployee`
        ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` 
        ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' 
        AND `tabEmployee`.payroll_category = '%s' AND 
        `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' AND `tabSalary Detail`.salary_component = 'EPF'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if lwf_amount['amount'] != None:
            total_lwf_amount += lwf_amount['amount']

        tds_amount = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` INNER JOIN 
        `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON 
        `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND 
        `tabEmployee`.payroll_category = '%s' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' 
        AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if tds_amount['amount'] != None:
            total_tds_amount += tds_amount['amount']

        salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS amount FROM `tabSalary Slip` INNER JOIN 
        `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON 
        `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND 
        `tabEmployee`.payroll_category = '%s' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' 
        AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if salary_advance['amount'] != None:
            total_salary_advance += salary_advance['amount']

        net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS amount FROM `tabSalary Slip` INNER JOIN `tabEmployee` 
        ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND 
        `tabEmployee`.payroll_category = '%s' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(c.name,args.from_date,args.to_date),as_dict=True)[0]
        if net_pay['amount'] != None:
            total_net_pay += net_pay['amount']
        
        ws.append([c.name,staff_count['count'],gross_amount['amount'] or '-',arrear_amount['amount'] or '-',
        birthday_amount['amount'] or '-',mobile_amount['amount'] or '-',
        pf_ammount['amount'] or '-',epf_amount['amount'] or '-',lwf_amount['amount'] or '-',
        tds_amount['amount'] or '-', salary_advance['amount'] or '-', net_pay['amount'] or '-'])

    ws.append(['Grand Total',total_staff_count,total_arrear_amount,total_gross_amount,
    total_birthday_amount,total_mobile_amount,total_pf_ammount,total_epf_amount,
    total_lwf_amount,total_tds_amount,total_salary_advance,total_net_pay]) 

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=12)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
    ws.merge_cells(start_row=2,start_column=5,end_row=2,end_column=12)

    align_center = Alignment(horizontal='center',vertical='center')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True,size=12)
        cell.alignment = align_center

    for cell in ws["2:2"]:
        cell.font = Font(bold=True,size=12)
        cell.alignment = align_center     

    ws['A1'].fill = PatternFill(fgColor="D7BDE2", fill_type = "solid")

    for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=12):
            for cell in header:
                cell.fill = PatternFill(fgColor='F5B7B1', fill_type = "solid")

    for header in ws.iter_rows(min_row=18, max_row=18, min_col=1, max_col=12):
            for cell in header:
                cell.fill = PatternFill(fgColor='85C1E9', fill_type = "solid")            

    border = Border(left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'))
    
    for rows in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=12):
        for cell in rows:
            cell.border = border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
