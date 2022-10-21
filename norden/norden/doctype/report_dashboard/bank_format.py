from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'Bank Format'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)

    ws.append(["Transaction Type","Bene Code","Bene Account Number","Amount","Bene Name","Customer Reference","Value Date","IFSC Code","Bene Bank","Bene Branch","Email ID","","","","","","","","","","","",""])
    salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date},['*']) 
    
    for ss in salary_slips:
        print(ss.employee_name)
        dt = datetime.strptime(args.from_date,'%Y-%m-%d')
        d = dt.strftime('%b')
        # emp = frappe.get_doc('Employee',{"employee":ss.employee},['*'])
        emp = frappe.get_doc('Employee',{"employee":ss.employee},['*'])
        ws.append(["",emp.first_name,ss.bank_account_no,ss.net_pay,ss.employee_name,d,ss.end_date,emp.ifsc_code,ss.bank_name,emp.bank_branch,emp.company_email])

        


        




    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'