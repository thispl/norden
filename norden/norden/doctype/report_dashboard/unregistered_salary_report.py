
from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.data import get_datetime
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Unregistered Salary Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.append([                       ])
    ws.append([                       ])
    ws.append([                       ])

    ws.append(["PYMT_PROD_TYPE_CODE","PYMT_MODE","DEBIT_ACC_NO","BNF_NAME","BENE_ACC_NO","BENE_IFSC","AMOUNT",
               "CREDIT_NAAR","PYMT_DATE","MOBILE_NUM","EMAIL_ID","REMARK", "REFERENCE_NO"])  

    ws.append(["Fixed Value:PAB_VENDOR","Allowed_values:FT,NEFT,RTGS,IMPS","Allowed values:12 digit ICICI Bank Account Number","Name of Beneficiary",
    "Account number of Beneficiary","IFSC code of the  beneficiary","Numeric value with decimal upto 2 places","30 Characters narration","Date of payment in format DD-MM-YYYY",
    "The mobile number of the Beneficiary is a 10-digit numeric value.If the Beneficiary alert is opted through SMS,alert will be sent to this mobile number",
    "E-mail ID of the Beneficiary.If the Beneficiary alert through e-mail or Payment advice is opted,an e-mail will be sent to this e-mail ID.Only 1 E-mail ID.",
    "Payment remarks:will be available in MIS" , "Addtional information which customer wants to add"])                                 

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=30)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=30)
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=30)
    
    align_center = Alignment(horizontal='right')
    
    if args.company:
        employee = frappe.get_all("Employee",{"status":"Active","company":args.company},['*'])

    

    for emp in employee:
        ss = frappe.get_value("Salary Slip",{"employee":emp.employee,"start_date":args.from_date,"end_date":args.to_date},["net_pay"])
        ps = frappe.get_value("Salary Slip",{"employee":emp.employee,"start_date":args.from_date,"end_date":args.to_date},["posting_date"])
        ed = frappe.get_value("Salary Slip",{"employee":emp.employee,"start_date":args.from_date,"end_date":args.to_date},["end_date"])
        

        ws.append([emp.payment_prod_type_code,emp.payment_mode,emp.debit_acc_no,emp.first_name,emp.bank_ac_no,emp.ifsc_code,ss,"Salary" " " + getdate(ed).strftime("%B-%y"),format_date(ps),emp.cell_number,emp.company_email,emp.remarks,emp.ref_no])

    align_center = Alignment(horizontal='right')
    



    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'



