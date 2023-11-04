import frappe
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)


@frappe.whitelist()
def download():
	filename = 'Summary Test Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	att_date = getdate(args.from_date)
	att_month = att_date.strftime("%B")
	att_year = att_date.year

	# List of salary components
	salary_components_all = ['Arrear','Birthday Bonus','Overtime','Internet','Increment Arrears','Referral Bonus','customer Entertainment','Travel Allowance','Others','incentive','Earned Leave Encashment']
	# Generate a comma-separated string of salary components
	components_string_all = "', '".join(salary_components_all)

	#total ncpl_gross_salary
	ncpl_staff_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	ncpl_profeessional_fee_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	ncpl_lease_rent_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	ncpl_intern_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	ncpl_staff_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as ads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	ncpl_profeessional_fee_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as proads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	ncpl_lease_rent_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as leaseads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	ncpl_intern_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as inads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	ncpl_staff_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	ncpl_profees_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	ncpl_lease_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	ncpl_intern_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
 
	if ncpl_staff_gross['gross'] is not None and ncpl_staff_ad['ads'] is not None:
		ncpl_subtract=ncpl_staff_gross['gross'] - ncpl_staff_ad['ads']
	else:
		ncpl_subtract = ncpl_staff_gross['gross'] or 0.0
	# ncpl_subtract = ncpl_staff_gross['gross'] - ncpl_staff_ad['ads']

	if ncpl_profeessional_fee_gross['gross'] is not None and ncpl_profeessional_fee_ad['proads'] is not None:
		ncpl_profeessional_fee_subtract=ncpl_profeessional_fee_gross['gross'] - ncpl_profeessional_fee_ad['proads']
	else:
		ncpl_profeessional_fee_subtract =  ncpl_profeessional_fee_gross['gross'] or 0.0
	# ncpl_profeessional_fee_subtract = ncpl_profeessional_fee_gross['gross'] - ncpl_profeessional_fee_ad['proads']

	if ncpl_lease_rent_gross['gross'] is not None and ncpl_lease_rent_ad['leaseads'] is not None:
		ncpl_lease_subtract=ncpl_lease_rent_gross['gross'] - ncpl_lease_rent_ad['leaseads']
	else:
		ncpl_lease_subtract = ncpl_lease_rent_gross['gross'] or 0.0
	# ncpl_lease_subtract = ncpl_lease_rent_gross['gross'] - ncpl_lease_rent_ad['leaseads']

	if ncpl_intern_gross['gross'] is not None and ncpl_intern_ad['inads'] is not None:
		ncpl_inad_subtract=ncpl_intern_gross['gross'] - ncpl_intern_ad['inads']
	else:
		ncpl_inad_subtract = ncpl_intern_gross['gross'] or 0.0
	# ncpl_inad_subtract = ncpl_intern_gross['gross'] - ncpl_intern_ad['inads']

	#total_nric_gross_salary
	nric_staff_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	marcom_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	technical_product_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_sea_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_uk_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_africa_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_freelance_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_intern_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	nric_contract_gross = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.gross_pay) AS gross FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_staff_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as ads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	marcom_gross_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as marads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	rnd_gross_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as rndads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	nric_sea_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as seaads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	uk_gross_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as ukads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0
	
	nric_africa_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as afads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	nric_freelance_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as freeads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	nric_intern_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as intads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	nric_contract_ad = frappe.db.sql("""SELECT SUM(`tabSalary Detail`.amount) as contads FROM `tabSalary Slip` 
	inner join `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	left join `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT'
	AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  
	and `tabSalary Detail`.salary_component in ('%s') """%(args.from_date,args.to_date,components_string_all),as_dict=True)[0] or 0.0

	nric_staff_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_marcom_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_technical_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_sea_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	nric_uk_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	nric_africa_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	nric_freelance_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	nric_intern_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
   
	nric_contract_lop = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.loss_of_pay) AS loss FROM `tabSalary Slip` 
	INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name 
	WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	
	if nric_staff_gross['gross'] is not None and nric_staff_ad['ads'] is not None:
		nric_subtract = nric_staff_gross['gross'] - nric_staff_ad['ads']
	else:
		nric_subtract = nric_staff_gross['gross'] or 0.0
	# nric_subtract = nric_staff_gross['gross'] - nric_staff_ad['ads']
	
	if marcom_gross['gross'] is not None and marcom_gross_ad['marads'] is not None:
		marcom_gross_ads = marcom_gross['gross'] - marcom_gross_ad['marads']
	else:
		marcom_gross_ads = marcom_gross['gross'] or 0.0
	# marcom_gross_ads = marcom_gross['gross'] -  marcom_gross_ad['marads']
	
	if technical_product_gross['gross'] is not None and rnd_gross_ad['rndads'] is not None:
		rnd_gross_ads = technical_product_gross['gross'] - rnd_gross_ad['rndads']
	else:
		rnd_gross_ads = technical_product_gross['gross'] or 0.0
	# rnd_gross_ads = technical_product_gross['gross'] - rnd_gross_ad['rndads']

	if nric_sea_gross['gross'] is not None and nric_sea_ad['seaads'] is not None:
		nric_sea_ads = nric_sea_gross['gross'] - nric_sea_ad['seaads']
	else:
		nric_sea_ads = nric_sea_gross['gross'] or 0.0
	# nric_sea_ads = nric_sea_gross['gross'] - nric_sea_ad['seaads']

	if nric_uk_gross['gross'] is not None and uk_gross_ad['ukads'] is not None:
		uk_gross_ads = nric_uk_gross['gross'] - uk_gross_ad['ukads']
	else:
		uk_gross_ads = nric_uk_gross['gross'] or 0.0
	# uk_gross_ads = nric_uk_gross['gross'] - uk_gross_ad['ukads']

	if nric_africa_gross['gross'] is not None and nric_africa_ad['afads'] is not None:
		nric_africa_ads = nric_africa_gross['gross'] - nric_africa_ad['afads']
	else:
		nric_africa_ads = nric_africa_gross['gross'] or 0.0
	# nric_africa_ads = nric_africa_gross['gross'] - nric_africa_ad['afads']

	if nric_freelance_gross['gross'] is not None and nric_freelance_ad['freeads'] is not None:
		nric_freelance_ads = nric_freelance_gross['gross'] - nric_freelance_ad['freeads']
	else:
		nric_freelance_ads = nric_freelance_gross['gross'] or 0.0
	# nric_freelance_ads = nric_freelance_gross['gross'] - nric_freelance_ad['freeads']
	
	if nric_intern_gross['gross'] is not None and nric_intern_ad['intads'] is not None:
		nric_intern_ads = nric_intern_gross['gross'] - nric_intern_ad['intads']
	else:
		nric_intern_ads = nric_intern_gross['gross'] or 0.0
	# nric_intern_ads = nric_intern_gross['gross'] - nric_intern_ad['intads']

	if nric_contract_gross['gross'] is not None and nric_contract_ad['contads'] is not None:
		nric_contract_ads = nric_contract_gross['gross'] - nric_contract_ad['contads']
	else:
		nric_contract_ads = nric_contract_gross['gross'] or 0.0
	# nric_contract_ads = nric_contract_gross['gross'] - nric_contract_ad['contads']

	#overall_ncpl_gross
	overall_ncpl_gross = sum([ncpl_staff_gross['gross'] or 0.0,ncpl_profeessional_fee_gross['gross'] or 0.0,ncpl_lease_rent_gross['gross'] or 0.0,ncpl_intern_gross['gross'] or 0.0])
	overall_ncpl_gross_without_ad = sum([ncpl_subtract,ncpl_profeessional_fee_subtract,ncpl_lease_subtract,ncpl_inad_subtract])
	
	#overall_nric_gross
	overall_nric_gross = sum([nric_staff_gross['gross'] or 0.0,marcom_gross['gross'] or 0.0,technical_product_gross['gross'] or 0.0,nric_sea_gross['gross'] or 0.0,nric_uk_gross['gross'] or 0.0,nric_africa_gross['gross'] or 0.0,nric_freelance_gross['gross'] or 0.0,nric_intern_gross['gross'] or 0.0,nric_contract_gross['gross'] or 0.0])
	overall_nric_gross_without_ad = sum([nric_subtract,marcom_gross_ads,rnd_gross_ads,nric_sea_ads,uk_gross_ads,nric_africa_ads,nric_freelance_ads,nric_intern_ads,nric_contract_ads])
	#grand_total_gross
	grand_total_gross = sum([overall_ncpl_gross,overall_nric_gross])
	grand_total_gross_without_ad =sum ([overall_ncpl_gross_without_ad,overall_nric_gross_without_ad])
	
	#total_ncpl_birthday_bonus
	ncpl_staff_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_profeessional_fee_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#total_nric_birthday_bonus
	nric_staff_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_freelance_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_birthday = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS birthday FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Birthday Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_birthday
	overall_ncpl_birthday = sum([ncpl_staff_birthday['birthday'] or 0.0,ncpl_profeessional_fee_birthday['birthday'] or 0.0,ncpl_lease_rent_birthday['birthday'] or 0.0,ncpl_intern_birthday['birthday'] or 0.0])
	#overall_total_nric_birthday
	overall_nric_birthday = sum([nric_staff_birthday['birthday'] or 0.0,marcom_birthday['birthday'] or 0.0,technical_product_birthday['birthday'] or 0.0,nric_sea_birthday['birthday'] or 0.0,nric_uk_birthday['birthday'] or 0.0,nric_africa_birthday['birthday'] or 0.0,nric_freelance_birthday['birthday'] or 0.0,nric_intern_birthday['birthday'] or 0.0,nric_contract_birthday['birthday']or 0.0])

	#grand_total_of_birthday
	grand_total_birthday = sum([overall_ncpl_birthday,overall_nric_birthday])

	#total_ncpl_internet
	ncpl_staff_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_internet
	nric_staff_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_internet = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS internet FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Internet'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_internet
	overall_ncpl_internet = sum([ncpl_staff_internet['internet'] or 0.0,ncpl_professional_fee_internet['internet'] or 0.0,ncpl_lease_rent_internet['internet'] or 0.0,ncpl_intern_internet['internet'] or 0.0])
	#overall_total_nric_internet
	overall_nric_internet = sum([nric_staff_internet['internet'] or 0.0,marcom_internet['internet'] or 0.0,technical_product_internet['internet'] or 0.0,nric_sea_internet['internet'] or 0.0,nric_uk_internet['internet'] or 0.0,nric_africa_internet['internet'] or 0.0,nric_freelance_internet['internet'] or 0.0,nric_intern_internet['internet'] or 0.0, nric_contract_internet['internet'] or 0.0])
	
	#grand_total_of_internet
	grand_total_of_internet = sum([overall_ncpl_internet,overall_nric_internet])

	# List of salary components
	salary_components = ['Arrear', 'Increment Arrears', 'Travel Allowance', 'Others']
	# Generate a comma-separated string of salary components
	components_string = "', '".join(salary_components)
	
	#total_ncpl_arrears
	ncpl_staff_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component IN ('%s')  """%(args.from_date,args.to_date,components_string),as_dict=True)[0]
	ncpl_professional_fee_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component IN ('%s')  """%(args.from_date,args.to_date,components_string),as_dict=True)[0] 
	
	#total_nric_arrears
	nric_staff_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component IN ('%s')  """%(args.from_date,args.to_date,components_string),as_dict=True)[0]
	nric_sea_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_arrears = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS arrears FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Arrear'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_arrears
	overall_ncpl_arrears = sum([ncpl_staff_arrears['arrears'] or 0.0,ncpl_professional_fee_arrears['arrears'] or 0.0,ncpl_lease_rent_arrears['arrears'] or 0.0,ncpl_intern_arrears['arrears'] or 0.0])
	#overall_total_ncpl_arrears
	overall_nric_arrears = sum([nric_staff_arrears['arrears'] or 0.0,marcom_salary_arrears['arrears'] or 0.0,technical_product_arrears['arrears'] or 0.0,nric_sea_arrears['arrears'] or 0.0,nric_uk_arrears['arrears'] or 0.0,nric_africa_arrears['arrears'] or 0.0,nric_freelance_arrears['arrears'] or 0.0,nric_intern_arrears['arrears'] or 0.0 , nric_contract_arrears['arrears'] or 0.0])

	#grand_total_of_arrears
	grand_total_of_arrears = sum([overall_ncpl_arrears,overall_nric_arrears])

	#total_ncpl_incentives
	ncpl_staff_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	ncpl_professional_fee_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'incentive'   """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'incentive'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	
	#total_nric_incentives
	nric_staff_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_incentives = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS incentives FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Incentive'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_incentives
	overall_ncpl_incentives = sum([ncpl_staff_incentives['incentives'] or 0.0,ncpl_professional_fee_incentives['incentives'] or 0.0,ncpl_lease_rent_incentives['incentives'] or 0.0,ncpl_intern_incentives['incentives'] or 0.0])
	#overall_total_ncpl_incentives
	overall_nric_incentives = sum([nric_staff_incentives['incentives'] or 0.0,marcom_salary_incentives['incentives'] or 0.0,technical_product_incentives['incentives'] or 0.0,nric_sea_incentives['incentives'] or 0.0,nric_uk_incentives['incentives'] or 0.0,nric_africa_incentives['incentives'] or 0.0,nric_freelance_incentives['incentives'] or 0.0,nric_intern_incentives['incentives'] or 0.0 , nric_contract_incentives['incentives'] or 0.0])

	#grand_total_of_incentives
	grand_total_of_incentives = sum([overall_ncpl_incentives,overall_nric_incentives])

	#total_ncpl_encashment
	ncpl_staff_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	
	#total_nric_encashment
	nric_staff_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_encashment = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS encash FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Earned Leave Encashment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_encash
	overall_ncpl_encashment = sum([ncpl_staff_encashment['encash'] or 0.0,ncpl_professional_fee_encashment['encash'] or 0.0,ncpl_lease_rent_encashment['encash'] or 0.0,ncpl_intern_encashment['encash'] or 0.0])
	#overall_total_ncpl_encash
	overall_nric_encashment = sum([nric_staff_encashment['encash'] or 0.0,marcom_salary_encashment['encash'] or 0.0,technical_product_encashment['encash'] or 0.0,nric_sea_encashment['encash'] or 0.0,nric_uk_encashment['encash'] or 0.0,nric_africa_encashment['encash'] or 0.0,nric_freelance_encashment['encash'] or 0.0,nric_intern_encashment['encash'] or 0.0 , nric_contract_encashment['encash'] or 0.0])

	#grand_total_of_encash
	grand_total_of_encashment = sum([overall_ncpl_encashment,overall_nric_encashment])

	#total_ncpl_overtime
	ncpl_staff_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_overtime
	nric_staff_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_overtime = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS overtime FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Overtime'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_overtime
	overall_ncpl_overtime = sum([ncpl_staff_overtime['overtime'] or 0.0,ncpl_professional_fee_overtime['overtime'] or 0.0,ncpl_lease_rent_overtime['overtime'] or 0.0,ncpl_intern_overtime['overtime'] or 0.0])
	#overall_total_nric_overtime
	overall_nric_overtime = sum([nric_staff_overtime['overtime'] or 0.0,marcom_salary_overtime['overtime'] or 0.0,technical_product_overtime['overtime'] or 0.0,nric_sea_overtime['overtime'] or 0.0,nric_uk_overtime['overtime'] or 0.0,nric_africa_overtime['overtime'] or 0.0,nric_freelance_overtime['overtime'] or 0.0,nric_intern_overtime['overtime'] or 0.0 , nric_contract_overtime['overtime'] or 0.0])

	#grand_total_of_overtime
	grand_total_of_overtime = sum([overall_ncpl_overtime,overall_nric_overtime])

	#total_ncpl_referral_bonus
	ncpl_staff_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_referral= frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_referral_bonus
	nric_staff_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_referral = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS referral FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Referral Bonus'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_referral_bonus
	overall_ncpl_referral = sum([ncpl_staff_referral['referral'] or 0.0,ncpl_professional_fee_referral['referral'] or 0.0,ncpl_lease_rent_referral['referral'] or 0.0,ncpl_intern_referral['referral'] or 0.0])
	#overall_total_nric_referral_bonus
	overall_nric_referral = sum([nric_staff_referral['referral'] or 0.0,marcom_salary_referral['referral'] or 0.0,technical_product_referral['referral'] or 0.0,nric_sea_referral['referral'] or 0.0,nric_uk_referral['referral'] or 0.0,nric_africa_referral['referral'] or 0.0,nric_freelance_referral['referral'] or 0.0,nric_intern_referral['referral'] or 0.0 , nric_contract_referral['referral'] or 0.0])
	
	#grand_total_of_referral_bonus
	grand_total_of_referral = sum([overall_ncpl_referral,overall_nric_referral])

	#total_ncpl_customer_enter
	ncpl_staff_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_customer_enter= frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_customer_enter= frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_customer_enter
	nric_staff_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_customer_enter = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS customer FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'customer Entertainment'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_customer_enter
	overall_ncpl_customer_enter = sum([ncpl_staff_customer_enter ['customer'] or 0.0,ncpl_professional_fee_customer_enter['customer'] or 0.0,ncpl_lease_rent_customer_enter['customer'] or 0.0,ncpl_intern_customer_enter['customer'] or 0.0])
	#overall_total_nric_customer_enter
	overall_nric_customer_enter= sum([nric_staff_customer_enter['customer'] or 0.0,marcom_salary_customer_enter['customer'] or 0.0,technical_product_customer_enter['customer'] or 0.0,nric_sea_customer_enter['customer'] or 0.0,nric_uk_customer_enter['customer'] or 0.0,nric_africa_customer_enter['customer'] or 0.0,nric_freelance_customer_enter['customer'] or 0.0,nric_intern_customer_enter['customer'] or 0.0 , nric_contract_customer_enter['customer'] or 0.0])
	
	#grand_total_of_customer_enter
	grand_total_of_customer_enter = sum([overall_ncpl_customer_enter,overall_nric_customer_enter])

	#total_ncpl_professional_tax
	ncpl_staff_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_profeesional_tax
	nric_staff_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_pt = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS pt FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Professional Tax'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_professional_tax
	overall_ncpl_pt = sum([ncpl_staff_pt['pt'] or 0.0,ncpl_professional_fee_pt['pt'] or 0.0,ncpl_lease_rent_pt['pt'] or 0.0,ncpl_intern_pt['pt'] or 0.0])
	#overall_total_nric_professional_tax
	overall_nric_pt = sum([nric_staff_pt['pt'] or 0.0,marcom_pt['pt'] or 0.0,technical_product_pt['pt'] or 0.0,nric_sea_pt['pt'] or 0.0,nric_uk_pt['pt'] or 0.0,nric_africa_pt['pt'] or 0.0,nric_freelance_pt['pt'] or 0.0,nric_intern_pt['pt'] or 0.0,nric_contract_pt['pt'] or 0.0])

	#grand_total_of_professional_tax
	grand_total_pt = sum([overall_ncpl_pt,overall_nric_pt])

	#total_ncpl_epf_dedcution
	ncpl_staff_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_epf_dedcution
	nric_staff_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_epf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS epf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Provident Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_epf_dedcution
	overall_ncpl_epf = sum([ncpl_staff_epf['epf'] or 0.0,ncpl_professional_fee_epf['epf'] or 0.0,ncpl_lease_rent_epf['epf'] or 0.0,ncpl_intern_epf['epf'] or 0.0])
	#overall_total_nric_epf_dedcution
	overall_nric_epf = sum([nric_staff_epf['epf'] or 0.0,marcom_epf['epf'] or 0.0,technical_product_epf['epf'] or 0.0,nric_sea_epf['epf'] or 0.0,nric_uk_epf['epf'] or 0.0,nric_africa_epf['epf'] or 0.0,nric_freelance_epf['epf'] or 0.0,nric_intern_epf['epf'] or 0.0,nric_contract_epf['epf']or 0.0])
	
	#grand_total_of_epf_dedcution
	grand_total_epf = sum([overall_ncpl_epf,overall_nric_epf])

	#total_ncpl_lwf
	ncpl_staff_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_lwf
	nric_staff_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_lwf = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS lwf FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Labour Welfare Fund'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_lwf
	overall_ncpl_lwf = sum([ncpl_staff_lwf['lwf'] or 0.0,ncpl_professional_fee_lwf['lwf'] or 0.0,ncpl_lease_rent_lwf['lwf'] or 0.0,ncpl_intern_lwf['lwf'] or 0.0])
	#overall_total_nric_epf_dedcution
	overall_nric_lwf = sum([nric_staff_lwf['lwf'] or 0.0,marcom_lwf['lwf'] or 0.0,technical_product_lwf['lwf'] or 0.0,nric_sea_lwf['lwf'] or 0.0,nric_uk_lwf['lwf'] or 0.0,nric_africa_lwf['lwf'] or 0.0,nric_freelance_lwf['lwf'] or 0.0,nric_intern_lwf['lwf'] or 0.0,nric_contract_lwf['lwf'] or 0.0])
	
	#grand_total_of_lwf
	grand_total_of_lwf = sum([overall_ncpl_lwf,overall_nric_lwf])

	#total_ncpl_tds
	ncpl_staff_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	
	#total_nric_tds
	nric_staff_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_tds = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS tds FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Tax Deducted Source'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_tds
	overall_ncpl_tds = sum([ncpl_staff_tds['tds'] or 0.0,ncpl_professional_fee_tds['tds'] or 0.0,ncpl_lease_rent_tds['tds'] or 0.0,ncpl_intern_tds['tds'] or 0.0])
	
	#overall_total_nric_tds
	overall_nric_tds = sum([nric_staff_tds['tds'] or 0.0,marcom_tds['tds'] or 0.0,technical_product_tds['tds'] or 0.0,nric_sea_tds['tds'] or 0.0,nric_uk_tds['tds'] or 0.0,nric_africa_tds['tds'] or 0.0,nric_freelance_tds['tds'] or 0.0,nric_intern_tds['tds'] or 0.0,nric_contract_tds['tds'] or 0.0])
	
	#grand_total_of_tds
	grand_total_of_tds = sum([overall_ncpl_tds,overall_nric_tds])
	
	#total_ncpl_esi
	ncpl_staff_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_esi
	nric_staff_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_esi = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS esi FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Employee State Insurance'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_esi
	overall_ncpl_esi = sum([ncpl_staff_esi['esi'] or 0.0,ncpl_professional_fee_esi['esi'] or 0.0,ncpl_lease_rent_esi['esi'] or 0.0,ncpl_intern_esi['esi'] or 0.0])
	
	#overall_total_nric_esi
	overall_nric_esi = sum([nric_staff_esi['esi'] or 0.0,marcom_esi['esi'] or 0.0,technical_product_esi['esi'] or 0.0,nric_sea_esi['esi'] or 0.0,nric_uk_esi['esi'] or 0.0,nric_africa_esi['esi'] or 0.0,nric_freelance_esi['esi'] or 0.0,nric_intern_esi['esi'] or 0.0,nric_contract_esi['esi'] or 0.0])
	
	#grand_total_of_esi
	grand_total_of_esi = sum([overall_ncpl_esi],overall_nric_esi)

	#total_ncpl_salary_advance
	ncpl_staff_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_salary_advance
	nric_staff_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_salary_advance = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS salary_advance FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Salary Advance Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]

	#overall_total_ncpl_salary-advance
	overall_ncpl_salary_advance = sum([ncpl_staff_salary_advance['salary_advance'] or 0.0,ncpl_professional_fee_salary_advance['salary_advance'] or 0.0,ncpl_lease_rent_salary_advance['salary_advance'] or 0.0,ncpl_intern_salary_advance['salary_advance'] or 0.0])
	#overall_total_nric_salary_advance
	overall_nric_salary_advance = sum([nric_staff_salary_advance['salary_advance'] or 0.0,marcom_salary_advance['salary_advance'] or 0.0,technical_product_salary_advance['salary_advance'] or 0.0,nric_sea_salary_advance['salary_advance'] or 0.0,nric_uk_salary_advance['salary_advance'] or 0.0,nric_africa_salary_advance['salary_advance'] or 0.0,nric_freelance_salary_advance['salary_advance'] or 0.0,nric_intern_salary_advance['salary_advance'] or 0.0,nric_contract_salary_advance['salary_advance'] or 0.0])
	
	#grand_total_of_salary_advance
	grand_total_of_salary_advance = sum([overall_ncpl_salary_advance,overall_nric_salary_advance])

	#total_ncpl_other_deduction
	ncpl_staff_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_professional_fee_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_lease_rent_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	ncpl_intern_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0] 

	#total_nric_other_deduction
	nric_staff_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	marcom_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	technical_product_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_sea_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_uk_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_africa_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0] 
	nric_freelance_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_intern_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	nric_contract_other_deduction = frappe.db.sql(""" SELECT SUM(`tabSalary Detail`.amount) AS other_deduction FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name INNER JOIN `tabSalary Detail` ON `tabSalary Slip`.name = `tabSalary Detail`.parent WHERE  `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'   AND `tabSalary Detail`.salary_component = 'Other Deduction'  """%(args.from_date,args.to_date),as_dict=True)[0]
	
	#overall_total_ncpl_other_deduction
	overall_ncpl_other_deduction = sum([ncpl_staff_other_deduction['other_deduction'] or 0.0,ncpl_professional_fee_other_deduction['other_deduction'] or 0.0,ncpl_lease_rent_other_deduction['other_deduction'] or 0.0,ncpl_intern_other_deduction['other_deduction']  or 0.0])
	#overall_total_nric_other_deduction
	overall_nric_other_deduction = sum([nric_staff_other_deduction['other_deduction'] or 0.0,marcom_other_deduction['other_deduction'] or 0.0,technical_product_other_deduction['other_deduction'] or 0.0,nric_sea_other_deduction['other_deduction'] or 0.0,nric_uk_other_deduction['other_deduction'] or 0.0,nric_africa_other_deduction['other_deduction'] or 0.0,nric_freelance_other_deduction['other_deduction'] or 0.0,nric_intern_other_deduction['other_deduction'] or 0.0,nric_contract_other_deduction['other_deduction'] or 0.0])
	
	#grand_total_of_other_deduction
	grand_total_of_other_deduction = sum([overall_ncpl_other_deduction,overall_nric_other_deduction])

	#overall_ncpl_loss_deduction
	overall_ncpl_loss_deduction = sum([ncpl_staff_lop['loss'] or 0.0,ncpl_profees_lop['loss'] or 0.0,ncpl_lease_lop['loss'] or 0.0,ncpl_intern_lop['loss'] or 0.0])
	
	#overall_nric_lop_deduction
	overall_nric_lop_deduction = sum([nric_staff_lop['loss'] or 0.0,nric_marcom_lop['loss'] or 0.0,nric_technical_lop['loss'] or 0.0,nric_sea_lop['loss'] or 0.0,nric_uk_lop['loss'] or 0.0,nric_africa_lop['loss'] or 0.0,nric_freelance_lop['loss'] or 0.0,nric_intern_lop['loss'] or 0.0,nric_contract_lop['loss'] or 0.0])
	
	#grand_total_of_lop_deduction
	grand_total_of_lop_deduction = sum([overall_ncpl_loss_deduction,overall_nric_lop_deduction])

	#total_ncpl_netpay
	ncpl_staff = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_professional_fee = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_lease_rent = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_intern = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
		
	#total_nric_netpay
	nric_staff = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	marcom_staff = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_technical_product = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_sea_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_uk_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_africa_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_freelance_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_intern_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_contract_net_pay = frappe.db.sql("""SELECT SUM(`tabSalary Slip`.net_pay) AS net_pay FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0

	#overall_ncpl_netpay
	overall_net_pay_ncpl = sum([ncpl_staff['net_pay'] or 0.0,ncpl_professional_fee['net_pay'] or 0.0,ncpl_lease_rent['net_pay'] or 0.0,ncpl_intern['net_pay'] or 0.0])

	#overall_nric_netpay
	overall_net_pay_nric = sum([nric_staff['net_pay'] or 0.0,marcom_staff['net_pay'] or 0.0,nric_technical_product['net_pay'] or 0.0,nric_sea_net_pay['net_pay'] or 0.0,nric_uk_net_pay['net_pay'] or 0.0,nric_africa_net_pay['net_pay'] or 0.0,nric_freelance_net_pay['net_pay'] or 0.0,nric_intern_net_pay['net_pay'] or 0.0,nric_contract_net_pay['net_pay'] or 0.0])

	#grand_total_of_netpay
	grand_total_netpay = sum([overall_net_pay_ncpl,overall_net_pay_nric])
				
	#Total No of Employees in NCPL Categories
	ncpl_staff_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL Staff' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_professional_fee_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL PROFESSIONAL FEE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_lease_rent_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL LEASE RENT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	ncpl_intern_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NCPL INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	total_ncpl_count = sum([ncpl_staff_count['count'],ncpl_professional_fee_count['count'],ncpl_lease_rent_count['count'],ncpl_intern_count['count']])
	
	#Total No of Employees in NRIC Categories
	nric_staff_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_marcom_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'MARCOM STAFF' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_technical_product_count = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'R & D / TECHNICAL / PRODUCT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_sea = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - SEA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_uk = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC -UK' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_africa = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC - AFRICA' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_freelance = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC FREELANCE' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s' """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0 
	nric_intern = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC INTERN' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	nric_contract = frappe.db.sql("""SELECT count(*) AS count FROM `tabSalary Slip` INNER JOIN `tabEmployee` ON `tabSalary Slip`.employee = `tabEmployee`.name WHERE `tabEmployee`.status = 'Active' AND `tabEmployee`.payroll_category = 'NRIC CONTRACT' AND `tabSalary Slip`.start_date BETWEEN '%s' AND '%s'  """%(args.from_date,args.to_date),as_dict=True)[0] or 0.0
	total_nric_count = sum([nric_staff_count['count'],nric_marcom_count['count'],nric_technical_product_count['count'],nric_sea['count'],
					nric_uk['count'],nric_africa['count'],nric_freelance['count'],nric_intern['count'],nric_contract['count']])

	#grand_total_all_employees_count
	grand_total_emp = sum([total_ncpl_count,total_nric_count])

	
	ws = wb.create_sheet(sheet_name, 0)  

	ws.append([str(att_month)+'-'+str(att_year)])
	
	ws.append(['','','','Additions','','','','','','','','Deductions','','','','','','','',''])
	
	ws.append(['Particulars','Number of Staffs','Gross','Arrears/Other Additions','Incentives','Earned Leave Encashment','Referral Bonus','Customer Entertainment','Overtime','Birthday Bonus','Mobile & Internet Allowance','Gross Pay W/o Arrear','Professional Tax','EPF Deduction','LWF','TDS-Salary','ESI','Salary Advance','Other Deduction','Loss of Pay','Net Amount'])
	
	ws.append(['NCPL Staff',ncpl_staff_count['count'],ncpl_staff_gross['gross'],ncpl_staff_arrears['arrears'],ncpl_staff_incentives['incentives'],ncpl_staff_encashment['encash'],ncpl_staff_referral['referral'],ncpl_staff_customer_enter['customer'],ncpl_staff_overtime['overtime'],ncpl_staff_birthday['birthday'],ncpl_staff_internet['internet'],ncpl_subtract,ncpl_staff_pt['pt'],ncpl_staff_epf['epf'],ncpl_staff_lwf['lwf'],ncpl_staff_tds['tds'],ncpl_staff_esi['esi'],ncpl_staff_salary_advance['salary_advance'],ncpl_staff_other_deduction['other_deduction'],ncpl_staff_lop['loss'],ncpl_staff['net_pay']])
	
	ws.append(['NCPL Professional Fee',ncpl_professional_fee_count['count'],ncpl_profeessional_fee_gross['gross'],ncpl_professional_fee_arrears['arrears'],ncpl_professional_fee_incentives['incentives'],ncpl_professional_fee_encashment['encash'],ncpl_professional_fee_referral['referral'],ncpl_professional_fee_customer_enter['customer'],ncpl_professional_fee_overtime['overtime'],ncpl_profeessional_fee_birthday['birthday'],ncpl_professional_fee_internet['internet'],ncpl_profeessional_fee_subtract,ncpl_professional_fee_pt['pt'],ncpl_professional_fee_epf['epf'],ncpl_professional_fee_lwf['lwf'],ncpl_professional_fee_tds['tds'],ncpl_professional_fee_esi['esi'],ncpl_professional_fee_salary_advance['salary_advance'],ncpl_professional_fee_other_deduction['other_deduction'],ncpl_profees_lop['loss'],ncpl_professional_fee['net_pay']])
	
	ws.append(['NCPL Lease Rent',ncpl_lease_rent_count['count'],ncpl_lease_rent_gross['gross'],ncpl_lease_rent_arrears['arrears'],ncpl_lease_rent_incentives['incentives'],ncpl_lease_rent_encashment['encash'],ncpl_lease_rent_referral['referral'],ncpl_lease_rent_customer_enter['customer'],ncpl_lease_rent_overtime['overtime'],ncpl_lease_rent_birthday['birthday'],ncpl_lease_rent_internet['internet'],ncpl_lease_subtract,ncpl_lease_rent_pt['pt'],ncpl_lease_rent_epf['epf'],ncpl_lease_rent_lwf['lwf'],ncpl_lease_rent_tds['tds'],ncpl_lease_rent_esi['esi'],ncpl_lease_rent_salary_advance['salary_advance'],ncpl_lease_rent_other_deduction['other_deduction'],ncpl_lease_lop['loss'],ncpl_lease_rent['net_pay']])
	
	ws.append(['NCPL Intern',ncpl_intern_count['count'],ncpl_intern_gross['gross'],ncpl_intern_arrears['arrears'],ncpl_intern_incentives['incentives'],ncpl_intern_encashment['encash'],ncpl_intern_referral['referral'],ncpl_intern_customer_enter['customer'],ncpl_intern_overtime['overtime'],ncpl_intern_birthday['birthday'],ncpl_intern_internet['internet'],ncpl_inad_subtract,ncpl_intern_pt['pt'],ncpl_intern_epf['epf'],ncpl_intern_lwf['lwf'],ncpl_intern_tds['tds'],ncpl_intern_esi['esi'],ncpl_intern_salary_advance['salary_advance'],ncpl_intern_other_deduction['other_deduction'],ncpl_intern_lop['loss'],ncpl_intern['net_pay']])
	
	ws.append(['Payroll NCPL',total_ncpl_count,overall_ncpl_gross,overall_ncpl_arrears,overall_ncpl_incentives,overall_ncpl_encashment,overall_ncpl_referral,overall_ncpl_customer_enter,overall_ncpl_overtime,overall_ncpl_birthday,overall_ncpl_internet,overall_ncpl_gross_without_ad,overall_ncpl_pt,overall_ncpl_epf,overall_ncpl_lwf,overall_ncpl_tds,overall_ncpl_esi,overall_ncpl_salary_advance,overall_ncpl_other_deduction,overall_ncpl_loss_deduction,overall_net_pay_ncpl]) 
	
	ws.append(['NRIC Staff',nric_staff_count['count'],nric_staff_gross['gross'],nric_staff_arrears['arrears'],nric_staff_incentives['incentives'],nric_staff_encashment['encash'],nric_staff_referral['referral'],nric_staff_customer_enter['customer'],nric_staff_overtime['overtime'],nric_staff_birthday['birthday'],nric_staff_internet['internet'],nric_subtract,nric_staff_pt['pt'],nric_staff_epf['epf'],nric_staff_lwf['lwf'],nric_staff_tds['tds'],nric_staff_esi['esi'],nric_staff_salary_advance['salary_advance'],nric_staff_other_deduction['other_deduction'],nric_staff_lop['loss'] ,nric_staff['net_pay']]) 
	
	ws.append(['Marcom Staff',nric_marcom_count['count'],marcom_gross['gross'],marcom_salary_arrears['arrears'],marcom_salary_incentives['incentives'],marcom_salary_encashment['encash'],marcom_salary_referral['referral'],marcom_salary_customer_enter['customer'],marcom_salary_overtime['overtime'],marcom_birthday['birthday'],marcom_internet['internet'],marcom_gross_ads,marcom_pt['pt'],marcom_epf['epf'],marcom_lwf['lwf'],marcom_tds['tds'],marcom_esi['esi'],marcom_salary_advance['salary_advance'],marcom_other_deduction['other_deduction'],nric_marcom_lop['loss'],marcom_staff['net_pay']]) 
	
	ws.append(['R&D/Technical/Product',nric_technical_product_count['count'],technical_product_gross['gross'],technical_product_arrears['arrears'],technical_product_incentives['incentives'],technical_product_encashment['encash'],technical_product_referral['referral'],technical_product_customer_enter['customer'],technical_product_overtime['overtime'],technical_product_birthday['birthday'],technical_product_internet['internet'],rnd_gross_ads,technical_product_pt['pt'],technical_product_epf['epf'],technical_product_lwf['lwf'],technical_product_tds['tds'],technical_product_esi['esi'],technical_product_salary_advance['salary_advance'],technical_product_other_deduction['other_deduction'],nric_technical_lop['loss'],nric_technical_product['net_pay']]) 
	
	ws.append(['NRIC-SEA',nric_sea['count'],nric_sea_gross['gross'],nric_sea_arrears['arrears'],nric_sea_incentives['incentives'],nric_sea_encashment['encash'],nric_sea_referral['referral'],nric_sea_customer_enter['customer'],nric_sea_overtime['overtime'],nric_sea_birthday['birthday'],nric_sea_internet['internet'],nric_sea_ads,nric_sea_pt['pt'],nric_sea_epf['epf'],nric_sea_lwf['lwf'],nric_sea_tds['tds'],nric_sea_esi['esi'],nric_sea_salary_advance['salary_advance'],nric_sea_other_deduction['other_deduction'],nric_sea_lop['loss'],nric_sea_net_pay['net_pay']]) 
	
	ws.append(['NRIC-UK',nric_uk['count'],nric_uk_gross['gross'],nric_uk_arrears['arrears'],nric_uk_incentives['incentives'],nric_uk_encashment['encash'],nric_uk_referral['referral'],nric_uk_customer_enter['customer'],nric_uk_overtime['overtime'],nric_uk_birthday['birthday'],nric_uk_internet['internet'],uk_gross_ads,nric_uk_pt['pt'],nric_uk_epf['epf'],nric_uk_lwf['lwf'],nric_uk_tds['tds'],nric_uk_esi['esi'],nric_uk_salary_advance['salary_advance'],nric_uk_other_deduction['other_deduction'],nric_uk_lop['loss'],nric_uk_net_pay['net_pay']]) 
	
	ws.append(['NRIC-Africa',nric_africa['count'],nric_africa_gross['gross'],nric_africa_arrears['arrears'],nric_africa_incentives['incentives'],nric_africa_encashment['encash'],nric_africa_referral['referral'],nric_africa_customer_enter['customer'],nric_africa_overtime['overtime'],nric_africa_birthday['birthday'],nric_africa_internet['internet'],nric_africa_ads,nric_africa_pt['pt'],nric_africa_epf['epf'],nric_africa_lwf['lwf'],nric_africa_tds['tds'],nric_africa_esi['esi'],nric_africa_salary_advance['salary_advance'],nric_africa_other_deduction['other_deduction'],nric_africa_lop['loss'],nric_africa_net_pay['net_pay']]) 
	
	ws.append(['NRIC Freelance',nric_freelance['count'],nric_freelance_gross['gross'],nric_freelance_arrears['arrears'],nric_freelance_incentives['incentives'],nric_freelance_encashment['encash'],nric_freelance_referral['referral'],nric_freelance_customer_enter['customer'],nric_freelance_overtime['overtime'],nric_freelance_birthday['birthday'],nric_freelance_internet['internet'],nric_freelance_ads,nric_freelance_pt['pt'],nric_freelance_epf['epf'],nric_freelance_lwf['lwf'],nric_freelance_tds['tds'],nric_freelance_esi['esi'],nric_freelance_salary_advance['salary_advance'],nric_freelance_other_deduction['other_deduction'],nric_freelance_lop['loss'],nric_freelance_net_pay['net_pay']]) 
	
	ws.append(['NRIC Intern',nric_intern['count'],nric_intern_gross['gross'],nric_intern_arrears['arrears'],nric_intern_incentives['incentives'],nric_intern_encashment['encash'],nric_intern_referral['referral'],nric_intern_customer_enter['customer'],nric_intern_overtime['overtime'],nric_intern_birthday['birthday'],nric_freelance_internet['internet'],nric_intern_ads,nric_intern_pt['pt'],nric_intern_epf['epf'],nric_intern_lwf['lwf'],nric_intern_tds['tds'],nric_intern_esi['esi'],nric_intern_salary_advance['salary_advance'],nric_intern_other_deduction['other_deduction'],nric_intern_lop['loss'],nric_intern_net_pay['net_pay']]) 

	ws.append(['NRIC Contract',nric_contract['count'],nric_contract_gross['gross'],nric_contract_arrears['arrears'],nric_contract_incentives['incentives'],nric_contract_encashment['encash'],nric_contract_referral['referral'],nric_contract_customer_enter['customer'],nric_contract_overtime['overtime'],nric_contract_birthday['birthday'],nric_contract_internet['internet'],nric_contract_ads,nric_contract_pt['pt'],nric_contract_epf['epf'],nric_contract_lwf['lwf'],nric_contract_tds['tds'],nric_contract_esi['esi'],nric_contract_salary_advance['salary_advance'],nric_contract_other_deduction['other_deduction'],nric_contract_lop['loss'] ,nric_contract_net_pay['net_pay']]) 

	ws.append(['Payroll NRIC',total_nric_count,overall_nric_gross,overall_nric_arrears,overall_nric_incentives,overall_nric_encashment,overall_nric_referral,overall_nric_customer_enter,overall_nric_overtime,overall_nric_birthday,overall_nric_internet,overall_nric_gross_without_ad,overall_nric_pt,overall_nric_epf,overall_nric_lwf,overall_nric_tds,overall_nric_esi,overall_nric_salary_advance,overall_nric_other_deduction,overall_nric_lop_deduction,overall_net_pay_nric]) 
	
	ws.append(['Payroll Electra','','','','','','','','','','','']) 
	
	ws.append(['Warehouse HouseKeeping Staff(NCPL)','','','','','','','','','','','']) 
	
	ws.append(['City Travel Allowance(NCPL)','','','','','','','','','','','']) 
	
	ws.append(['Grand Total',grand_total_emp,grand_total_gross,grand_total_of_arrears,grand_total_of_incentives,grand_total_of_encashment,grand_total_of_referral,grand_total_of_customer_enter,grand_total_of_overtime,grand_total_birthday,grand_total_of_internet,grand_total_gross_without_ad,grand_total_pt,grand_total_epf,grand_total_of_lwf,grand_total_of_tds,grand_total_of_esi,grand_total_of_salary_advance,grand_total_of_other_deduction,grand_total_of_lop_deduction,grand_total_netpay]) 

	ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=21)
	ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=3)
	ws.merge_cells(start_row=2,start_column=4,end_row=2,end_column=11)
	ws.merge_cells(start_row=2,start_column=12,end_row=2,end_column=20)
	# ws.merge_cells(start_row=2,start_column=12,end_row=2,end_column=20)

	align_center = Alignment(horizontal='center',vertical='center')

	for cell in ws["1:1"]:
		cell.font = Font(bold=True,size=14)
		cell.alignment = align_center

	for cell in ws["2:2"]:
		cell.font = Font(bold=True,size=14)
		cell.alignment = align_center     

	ws['A1'].fill = PatternFill(fgColor="D7BDE2", fill_type = "solid")

	for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='F5B7B1', fill_type = "solid")

	for header in ws.iter_rows(min_row=8, max_row=8, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='D7BDE2', fill_type = "solid")

	for header in ws.iter_rows(min_row=18, max_row=18, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='85C1E9', fill_type = "solid")            

	for header in ws.iter_rows(min_row=19, max_row=19, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='85C1E9', fill_type = "solid")

	for header in ws.iter_rows(min_row=20, max_row=20, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='D7BDE2', fill_type = "solid")

	for header in ws.iter_rows(min_row=21, max_row=21, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='D7BDE2', fill_type = "solid")

	for header in ws.iter_rows(min_row=22, max_row=22, min_col=1, max_col=21):
			for cell in header:
				cell.fill = PatternFill(fgColor='2471A3', fill_type = "solid")    

	border = Border(left=Side(border_style='thin', color='000000'),
			 right=Side(border_style='thin', color='000000'),
			 top=Side(border_style='thin', color='000000'),
			 bottom=Side(border_style='thin', color='000000'))
	
	for rows in ws.iter_rows(min_row=1, max_row=22, min_col=1, max_col=21):
		for cell in rows:
			cell.border = border

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'
