from curses import is_term_resized
import email
import frappe
import requests
from frappe.utils.data import format_date, today 
from frappe import _
import json
from frappe.utils.background_jobs import enqueue
from norden.custom import employee
from frappe.utils import flt
from dateutil import relativedelta
from datetime import date, datetime
from frappe.utils.csvutils import read_csv_content
from frappe.utils.file_manager import get_file
from erpnext.setup.utils import get_exchange_rate
from frappe.utils import (
	add_days,
	add_months,
	add_years,
	cint,
	cstr,
	date_diff,
	flt,
	formatdate,
	get_last_day,
	get_timestamp,
	getdate,
	nowdate,
)
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types



@frappe.whitelist()
def update_company_currency(account_currency,company_currency):
	return get_exchange_rate(account_currency,company_currency)

@frappe.whitelist()
def update_unit_price_document_currency():
	qi = frappe.db.sql("""select unit_price,unit_price_document_currency from `tabQuotation Item`  """)
	print(qi)
	
# @frappe.whitelist()
# def getstock_detail(item_details,company):
# 	item_details = json.loads(item_details)
# 	frappe.errprint(item_details)
# 	data = ''
# 	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
# 	data += '</table>'
# 	for j in item_details:
# 		country = frappe.get_value("Company",{"name":company},["country"])

# 		warehouse_stock = frappe.db.sql("""
# 		select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
# 		""" % (country,j["item_code"]),as_dict=True)[0]

# 		if not warehouse_stock["qty"]:
# 			warehouse_stock["qty"] = 0
# 		purchase_order = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
# 				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
# 				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(j["item_code"]),as_dict=True)[0] or 0 
# 		if not purchase_order["qty"]:
# 			purchase_order["qty"] = 0
# 		purchase_receipt = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.qty) as qty from `tabPurchase Receipt`
# 				left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
# 				where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt`.docstatus = 1 """%(j["item_code"]),as_dict=True)[0] or 0 
# 		if not purchase_receipt["qty"]:
# 			purchase_receipt["qty"] = 0
# 		in_transit = purchase_order["qty"] - purchase_receipt["qty"]
# 		total = warehouse_stock["qty"] + in_transit

# 		stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
# 		where item_code = '%s' """%(j["item_code"]),as_dict=True)

# 		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
# 		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
# 		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (j["item_code"]), as_dict=True)
	
# 		sale = frappe.db.sql("""select `tabSales Order Item`.item_code as item_code,`tabSales Order Item`.item_name as item_name,sum(`tabSales Order Item`.qty) as qty,`tabSales Order Item`.rate as rate,`tabSales Order`.name as so from `tabSales Order`
# 		left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
# 		where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 0 order by rate asc limit 1""" % (j["item_code"]), as_dict=True)
		
# 		i = 0
# 		for po in pos:
# 			for so in sale:
				
# 				if pos:
# 					frappe.errprint(po.qty)
# 					data += '<table class="table table-bordered">'
# 					data += '<tr>'
# 					data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM CODE</b><center></td>'
# 					data += '<td colspan=1 style="width:33%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM NAME</b><center></td>'
# 					data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>STOCK</b><center></td>'

# 					for stock in stocks:
# 						if stock.actual_qty > 0:
# 							wh = stock.warehouse
# 							x = wh.split('- ')
# 							data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>%s</b><center></td>'%(x[-1])
# 					data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PENDING TO RECEIVE</b><center></td>'
# 					# data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>AGAINST PO</b><center></td>'
# 					data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PENDING TO SELL</b><center></td>'
# 					data += '</tr>'
					
					
					
# 					data +='<tr>'
# 					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(j["item_code"])
# 					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(j["item_name"])
# 					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(warehouse_stock['qty'] or 0)
# 					for stock in stocks:
# 						if stock.actual_qty > 0:
# 							data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(stock.actual_qty)
# 					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(in_transit or 0)
# 					# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(po.qty or 0)
# 					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'%(so.qty or 0)
# 					data += '</tr>'
# 				i += 1
# 			data += '</table>'
			
# 		#     data+='<tr><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s<center><center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s</center></td><td colspan=1 style="border: 1px solid black;font-size:12px;"><center>%s</center></td></tr>'%(j["item_code"],j["item_name"],warehouse_stock["qty"], in_transit,total)
		
# 		# data += '</table>'
# 	return data

@frappe.whitelist()
def getstock_detail(item_details, company):
	item_details = json.loads(item_details)
	frappe.errprint(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS</b></center></h4>'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM CODE</b></center></td>'
	data += '<td colspan=1 style="width:33%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM NAME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>STOCK</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>NSPL</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>NCME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>NCUL</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>SNTL</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>NCPL</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>NRIC</b></center></td>'
	data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>IN TRANSIT</b></center></td>'
	data += '<td colspan=1 style="padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PENDING TO SELL</b></center></td>'
	warehouses = []  # Create a list to store warehouse names

	for j in item_details:
		country = frappe.get_value("Company", {"name": company}, ["country"])

		warehouse_stock = frappe.db.sql("""
		select sum(b.actual_qty) as qty from `tabBin` b join `tabWarehouse` wh on wh.name = b.warehouse join `tabCompany` c on c.name = wh.company where c.country = '%s' and b.item_code = '%s'
		""" % (country, j["item_code"]), as_dict=True)[0]

		if not warehouse_stock["qty"]:
			warehouse_stock["qty"] = 0
		purchase_order = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """ % (
			j["item_code"]), as_dict=True)[0] or 0
		if not purchase_order["qty"]:
			purchase_order["qty"] = 0
		purchase_receipt = frappe.db.sql("""select sum(`tabPurchase Receipt Item`.qty) as qty from `tabPurchase Receipt`
				left join `tabPurchase Receipt Item` on `tabPurchase Receipt`.name = `tabPurchase Receipt Item`.parent
				where `tabPurchase Receipt Item`.item_code = '%s' and `tabPurchase Receipt`.docstatus = 1 """ % (
			j["item_code"]), as_dict=True)[0] or 0
		if not purchase_receipt["qty"]:
			purchase_receipt["qty"] = 0
		in_transit = purchase_order["qty"] - purchase_receipt["qty"]
		total = warehouse_stock["qty"] + in_transit

		

		stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' """ % (j["item_code"]), as_dict=True)

		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,sum(`tabPurchase Order Item`.qty) as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order`.name as po from `tabPurchase Order`
				left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
				where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by rate asc limit 1""" % (
			j["item_code"]), as_dict=True)

		sale = frappe.db.sql("""select `tabSales Order Item`.item_code as item_code,`tabSales Order Item`.item_name as item_name,sum(`tabSales Order Item`.qty) as qty,`tabSales Order Item`.rate as rate,`tabSales Order`.name as so from `tabSales Order`
				left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
				where `tabSales Order Item`.item_code = '%s' and `tabSales Order`.docstatus = 0 order by rate asc limit 1""" % (
			j["item_code"]), as_dict=True)

		for po in pos:
			for so in sale:
				data += '<tr>'
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_code"])
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_name"])
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (warehouse_stock['qty'] or 0)
				
				compa = ["Norden Singapore PTE LTD","Norden Communication Middle East FZE","Norden Communication UK Limited","Sparcom Ningbo Telecom Ltd","Norden Communication Pvt Ltd","Norden Research and Innovation Centre (OPC) Pvt. Ltd"]

				for co in compa:
					st = 0
					ware = frappe.db.get_list("Warehouse",{"company":co},['name'])
					for w in ware:
						sto = frappe.db.get_value("Bin",{"item_code":j["item_code"],"warehouse":w.name},['actual_qty'])
						if not sto:
							sto = 0
						st += sto
					data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(st)
				# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' 
				# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'
				# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'
				# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'
				# data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>'

				# for stock in stocks:
				# 	wh = stock.warehouse
				# 	x = wh.split('- ')
				# 	warehouse_name = x[-1]
				# 	if warehouse_name not in warehouses:
				# 		warehouses.append(warehouse_name)  # Add warehouse name to the list
				# 	data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (stock.actual_qty)
					

				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (in_transit or 0)
				data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (so.qty or 0)
				data += '</tr>'

	data += '<tr>'
	
	data += '</table>'
	return data


@frappe.whitelist()
def item_default_wh(doc,method):
	item_group = frappe.get_value('Item',doc.item_code,'item_group')
	general_services = ['General Service','Service','Office Stationary','Assets','Stationary']
	if item_group not in general_services:
		item_default_set = frappe.get_value('Item',doc.item_code,'item_default_set')
		if not item_default_set:
			companies = [
			{
				"company": "Norden Communication Pvt Ltd",
				"buying_cost_center": "Main - NCPL",
				"selling_cost_center": "Main - NCPL",
				"expense_account": "Cost of Goods Sold - NCPL",
				"income_account": "Sales - NCPL"
			},
			{
				"company": "Norden Communication Middle East FZE",
				"buying_cost_center": "Main - NCME",
				"selling_cost_center": "Main - NCME",
				"expense_account": "Cost of Goods Sold - NCME",
				"income_account": "Sales - NCME"
			},
			{
				"company": "Norden Research and Innovation Centre (OPC) Pvt. Ltd",
				"buying_cost_center": "Main - NRIC",
				"selling_cost_center": "Main - NRIC",
				"expense_account": "Cost of Goods Sold - NRIC",
				"income_account": "Sales - NRIC"
			},
			{
				"company": "Norden Communication UK Limited",
				"buying_cost_center": "Main - NCUL",
				"selling_cost_center": "Main - NCUL",
				"expense_account": "Cost of Goods Sold - NCUL",
				"income_account": "Sales - NCUL"
			},
			{
				"company": "Sparcom Ningbo Telecom Ltd",
				"buying_cost_center": "Main - SNTL",
				"selling_cost_center": "Main - SNTL",
				"expense_account": "Cost of Goods Sold - SNTL",
				"income_account": "Sales - SNTL"
			},
			{
				"company": "Norden Singapore PTE LTD",
				"buying_cost_center": "Main - NSPL",
				"selling_cost_center": "Main - NSPL",
				"expense_account": "Cost of Goods Sold - NSPL",
				"income_account": "Sales - NSPL"
			},
			{
				"company": "Norden Communication India",
				"buying_cost_center": "Corporate - NC",
				"selling_cost_center": "Corporate - NC",
				"expense_account": "Cost of Goods Sold - NC",
				"income_account": "Sales - NC"
			}
						]
			for company in companies:
				itemid = ''
				item_default = frappe.db.exists('Item Default',{'parent':doc.item_code,'company':company['company']},'parent')
				if not item_default:
					itemid = frappe.get_doc("Item",doc.item_code)
					itemid.item_default_set = 1
					itemid.append('item_defaults',{
						'company':company['company'],
						'buying_cost_center':company['buying_cost_center'],
						'selling_cost_center':company['selling_cost_center'],
						'expense_account':company['expense_account'],
						'income_account':company['income_account'],
					})
					itemid.save(ignore_permissions=True)
				else:
					frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'buying_cost_center',company['buying_cost_center'])
					frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'selling_cost_center',company['selling_cost_center'])
					frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'expense_account',company['expense_account'])
					frappe.db.set_value('Item Default',{'parent':itemid.name,'company':company['company']},'income_account',company['income_account'])
			frappe.db.set_value('Item',doc.item_code,"item_default_set",1)



@frappe.whitelist()
def get_series(company,doctype):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype,'with_tax':0},'series')
	return company_series


@frappe.whitelist()
def get_serie():
	company = "Norden Communication Pvt Ltd"
	doctype = "Sales Invoice"
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype,'with_tax':0},'series')
	return company_series

@frappe.whitelist()
def get_series_with_tax(company,doctype):
	company_series = frappe.db.get_value("Company Series",{'company':company,'document_type':doctype,'with_tax':1},'series')
	return company_series
	
@frappe.whitelist()
def email_salary_slip():
	ss = frappe.get_all("Salary Slip",{'start_date':'2023-10-01'},['employee','name'])
	for s in ss:
		doc = frappe.get_doc("Salary Slip",s['name'])
		receiver = frappe.db.get_value("Employee", doc.employee, "company_email")
		payroll_settings = frappe.get_single("Payroll Settings")
		message = "Please Find the attachment"
		password = None

		if receiver:
			email_args = {
				"sender": "hrd@nordencommunication.com",
				"recipients":receiver,
				"message": _(message),
				"subject": 'Salary Slip - from {0} to {1}'.format(doc.start_date, doc.end_date),
				"attachments": [frappe.attach_print(doc.doctype, doc.name, file_name=doc.name, password=password)],
				"reference_doctype": doc.doctype,
				"reference_name": doc.name
				}
			if not frappe.flags.in_test:
				print('yes')
				enqueue(method=frappe.sendmail, queue='short', timeout=300, is_async=True, **email_args)
			else:
				print('No')
				frappe.sendmail(**email_args)
		else:
			print(_("{0}: Employee email not found, hence email not sent").format(doc.employee_name))

@frappe.whitelist()
def email_notification():
	employee = frappe.db.get_all('Employee',{'status':'Active','Company':'Norden Communication Pvt Ltd'},['*'])
	for emp in employee:
		if emp.company_email:
			frappe.sendmail(
				recipients = emp.company_email,
				# 'api@groupteampro.com','hrd@nordencommunication.com','aiswarya@nordencommunication.com','kareem@nordencommunication.com'],
				subject = ' ERP Notification- Travel Request and Expense Claim Online',
				message = """<p style='font-size:15px'>Dear All,</p><br>
						<p>On behalf of our Management, HRD is very pleased to introduce you to an online Travel Request and Expense Claim process through our new ERP system.</p><br>
						<p>Effective today, all Travel requests and expense Claims can be processed online (submission and approval), the system will allow you to add all your claims online, and also allow you to upload all relevant bills to support your claims. The approvals and the payment will be processed within two weeks of submission; however, you will be able to monitor the status and the stage of your bill claim online.</p><br>
						<p>Submit your claim through ERP access under Travel request & Expense Claim, and HR & Admin Expert- Ms. Aiswarya will be able to clarify the process in case you need support.</p><br>
						<p>Kindly find your user id and Password, you may connect the below hyperlink for a demo of the process for self-guidelines.</p><br>
						<table border="1">
						<tr><td>Employee ID</td><td>%s</td></tr>
						<tr><td>Employee Name</td><td>%s</td></tr>
						<tr><td>User Login</td><td>%s</td></tr>
						<tr><td>Password</td><td>Norden@1234</td></tr>
						<tr><td>ERP WEBSITE LINK</td><td><a href="https://erp.nordencommunication.com">erp.nordencommunication.com</a></td></tr>
						<tr><td><b>Travel Request WorkFlow<b></td><td><a href="https://screen-recorder-bucket.s3.ap-south-1.amazonaws.com/ScreenRecorder_2022-10-18_dfdd4ccb-b0e7-4bef-bb9a-2283f328c20c.mp4">Travel Request</a></td></tr>
						<tr><td><b>Expense Claim WorkFlow<b></td><td><a href="https://watch.screencastify.com/v/WfShLegErakwNcUqJ3X8">Expense claim</a></td></tr>
						</table><br>
						<p>Regards,<br>
						HRM</p><br>
						"""%(emp.employee_number,emp.first_name,emp.user_id)
			)
		else:
			message = ('Company Email Not in Employee %s'%(emp.name))   
			frappe.log_error('Email Notification',message) 
			



@frappe.whitelist()
def email_probation_emp():
	company_names = ["Norden Communication Pvt Ltd", "Norden Research and Innovation Centre (OPC) Pvt. Ltd"]
	category = ["NCPL INTERN" , "NRIC INTERN"]
	data = ''
	employee = frappe.get_all('Employee', {'status': 'Active', 'company': ['in' , company_names], 'payroll_category' : ['in' , category]}, [
							  "name", "employee_name", "department","date_of_joining"])
	data += 'Kindly Find the List of Employees going to complete their Internship<br><br><table class="table table-bordered">'
	data += '<table class="table table-bordered"><tr><th>Employee ID</th><th>Employee Name</th><th>Department</th><th>Date of Joining</th><th>Internship End Date</th></tr>'
	for emp in employee:
		intern_end_date = add_months(emp.date_of_joining, +6)
		data += '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' % (
				emp.name, emp.employee_name, emp.department, format_date(emp.date_of_joining),format_date(intern_end_date))
	data += '</table>'
	frappe.sendmail(
		recipients=['gayathri.r@groupteampro.com'],
		subject=('Reg: Employee Intership End Related'),
		header=('Dear HR'),
		message=data
	)

#Material Request Showing Available Stock and Previous Purchase Orders
#Stock Availablity HTML Table View
@frappe.whitelist()
def stock_available(item_details):
	item_detail = json.loads(item_details)
	data =''
	data +='<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=17><center>Stock Available</center></th></tr>'
	data +='<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Description</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NCFME</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NCUL</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NCPL</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NC</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse 5</b></td><td style="padding:1px;border: 1px solid black"><b>Total</b></td>'
	for s in item_detail:
		total_actual_qty = 0
		actual_qty = []
		description = frappe.get_value('Item',s['item_code'],'description')
		ware_house = ['Stores - NCFME','Stores - NCUL','Stores - NSPL','Stores - NC','Stores - NCI']
		for w in ware_house:
			stock_qty = frappe.db.sql(""" select actual_qty from `tabBin` where item_code = '%s' and warehouse = '%s' """%(s['item_code'],w),as_dict=1)
			if stock_qty:
				qty = stock_qty[0]['actual_qty']
				actual_qty.append(qty)
				total_actual_qty += qty
			else:
				actual_qty.append(0)
	
		data += '<tr><td style="padding: 1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td>'%(s['item_code'] or '',description or '',actual_qty[0],actual_qty[1],actual_qty[2],actual_qty[3],actual_qty[4],total_actual_qty)
	return data
# Previous Purchase Orders
# @frappe.whitelist()
# def previous_purchase_orders(item_details):
#    item_detail = json.loads(item_details)
#    data =''
#    data +='<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=17><center>Previous Purchase Orders</center></th></tr>'
#    data +='<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Description</b></td><td style="padding:1px;border: 1px solid black"><b>Supplier</b></td>'
# #    <td style="padding:1px;border: 1px solid black"><b>Warehouse NCUL</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NCPL</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse NC</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse 5</b></td><td style="padding:1px;border: 1px solid black"><b>Total</b></td>'
#    for s in item_detail:
# #        total_actual_qty = 0
# #        actual_qty = []
#        description = frappe.get_value('Item',s['item_code'],'description')
#     #    supplier = 
# #        ware_house = ['Stores - NCFME','Stores - NCUL','Stores - NSPL','Stores - NC','Stores - NCI']
# #        for w in ware_house:
# #            stock_qty = frappe.db.sql(""" select actual_qty from `tabBin` where item_code = '%s' and warehouse = '%s' """%(s['item_code'],w),as_dict=1)
# #            if stock_qty:
# #                qty = stock_qty[0]['actual_qty']
# #                actual_qty.append(qty)
# #                total_actual_qty += qty
# #            else:
# #                actual_qty.append(0)
 
# #        data += '<tr><td style="padding: 1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td><td style="padding:1px;border: 1px solid black"><b>%s</b></td>'%(s['item_code'] or '',description or '',actual_qty[0],actual_qty[1],actual_qty[2],actual_qty[3],actual_qty[4],total_actual_qty)
#    return data
 
#Quotation in Child Table Item Select will shown the stock items of the selected items 
@frappe.whitelist()
def stock_popup_table(item_code):
	item = frappe.db.get_value('Item',{'item_code':item_code},['item_code'])
	data = ''
	# stocks = frappe.db.sql(""" select actual_qty, warehouse """)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=6><center>Stock Availability</center></th></tr>'
	return data

#Purchase order to select child table to fetch the previous stock
@frappe.whitelist()
def purchase_order_items(item_code):
	item = frappe.db.get_value('Item',{'item_code':item_code},['item_code'])
	data = ''
	po = frappe.db.sql(""" select `tabPurchase Order Item` .item_code as item_code, `tabPurchase Order Item` .item_name as item_name,
						`tabPurchase Order Item` .supplier as supplier,`tabPurchase Order Item` .qty as qty,`tabPurchase Order`.transaction_date as date,
						`tabPurchase Order`.name as po from `tabPurchase Order` left join `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus !=2 """%(item_code),as_dict=True)
	
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=6><center>Stock Availability</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Item Name</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse</b></td><td style="padding:1px;border: 1px solid black"><b>QTY</b></td><td style="padding:1px;border: 1px solid black"><b>UOM</b></td></tr>'
	i = 0
	data +='</table>'
	return data

#Material Request to select table to fetch the Stock Availablity
@frappe.whitelist()
def stock_popup(item_code):
	item = frappe.get_value('Item',{'item_code':item_code},'item_code')
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin
		where item_code = '%s' """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black" colspan=6><center>Stock Availability</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black"><b>Item Code</b></td><td style="padding:1px;border: 1px solid black"><b>Item Name</b></td><td style="padding:1px;border: 1px solid black"><b>Warehouse</b></td><td style="padding:1px;border: 1px solid black"><b>QTY</b></td><td style="padding:1px;border: 1px solid black"><b>UOM</b></td></tr>'
	i = 0
	for stock in stocks:
		if stock.actual_qty > 0:
			data += '<tr><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td><td style="padding:1px;border: 1px solid black">%s</td></tr>'%(item,frappe.db.get_value('Item',item,'item_name'),stock.warehouse,stock.actual_qty,stock.stock_uom)
			i += 1
	data += '</table>'
	if i > 0:
		return data

@frappe.whitelist()
def get_hsn():
	item = frappe.get_value('Item',{'item_code':item_code},'item_code')
	return "hi"

def get_exc_rate():
	from_currency ='USD'
	to_currency = 'AED'
	exc_rate = get_exchange_rate(from_currency,to_currency)
	print(exc_rate)

@frappe.whitelist()
def set_territory():
	quotation = frappe.get_all('Quotation',{"status":"Ordered"},["*"])
	for i in quotation:
		territory = frappe.get_value('Customer',{'name':i.customer_name},['territory'])
		frappe.db.set_value('Quotation',i.name,'territory',territory)


@frappe.whitelist()
def get_user_id_travel_request(reports_to):
	employee = frappe.db.get_value('Employee',{'status':'Active','employee_number':reports_to},['employee_name','user_id'])
	return employee

@frappe.whitelist()
def get_user_id():
	user_login =  frappe.db.get_value('Employee',{'user_id':frappe.session.user},['user_id'])
	return user_login



@frappe.whitelist()
def get_empl(doc,method):
	salary_slip = frappe.db.sql("""select * from `tabSalary Slip` where status = 'Draft' and start_date between '2022-10-01' and '2022-10-31' """,as_dict=1)
	for esi in salary_slip:
		get=frappe.get_doc('Salary Slip',{'name':esi.name})
		frappe.errprint(get.earnings.salary_component)

@frappe.whitelist()
def get_customer_det(doc,method):
	# frappe.errprint("hi")
	if doc.first_name:
		cus = frappe.new_doc("Contact")
		cus.first_name = doc.first_name
		cus.last_name = doc.last_name
		cus.designation = doc.designation
		cus.status = doc.status
		cus.append("links", {
			"link_doctype": "Customer",
			"link_name":doc.name
			})
		cus.flags.ignore_mandatory = True    
		cus.save(ignore_permissions= True)

@frappe.whitelist()
def get_address_det(doc,method):
	# frappe.errprint("hi")
	if doc.address_title:
		addr = frappe.new_doc("Address")
		addr.address_title = doc.address_title
		addr.address_line1 = doc.address_line1
		addr.address_line2 = doc.address_line2
		addr.city = doc.city
		addr.country = doc.country
		addr.email_idl = doc.email_id
		addr.phone = doc.phone
		addr.append("links", {
			"link_doctype": "Customer",
			"link_name":doc.name
			})
		addr.flags.ignore_mandatory = True    
		addr.save(ignore_permissions= True)

@frappe.whitelist()
def create_sample_inspection(sample,item,po,pr,name):
	# s = int(sample) - 1
	it = frappe.get_value("Item",{"name":item},["inspection_template"])
	for i in range(int(sample)):
		doc = frappe.new_doc("Inspection Sample")
		doc.item = item
		doc.po_number = po
		doc.pr_number = pr
		doc.inspection_template = it
		doc.item_inspection = name
		doc.save(ignore_permissions = True)

@frappe.whitelist()
def create_sample_inspection_from_lr(sample,item,po,lr,name):
	# s = int(sample) - 1
	it = frappe.get_value("Item",{"name":item},["factory_inspection_template"])
	for i in range(int(sample)):
		doc = frappe.new_doc("Inspection Sample")
		doc.item = item
		doc.po_number = po
		doc.logistics_request_number = lr
		doc.inspection_template = it
		doc.item_inspection = name
		doc.save(ignore_permissions = True)

@frappe.whitelist()
def create_sample_inspection_from_dn(sample,item,so,dn,name):
	# s = int(sample) - 1
	it = frappe.get_value("Item",{"name":item},["inspection_template"])
	for i in range(int(sample)):
		doc = frappe.new_doc("Inspection Sample")
		doc.item = item
		doc.so_number = so
		doc.dn_number = dn
		doc.inspection_template = it
		doc.item_inspection = name
		doc.save(ignore_permissions = True)

@frappe.whitelist()
def update_quotation_cluster(import_file):
	filepath = get_file(import_file)
	pps = read_csv_content(filepath[1])
	for pp in pps:
		if frappe.db.exists('Sales Order',{'name':pp[1]}):
			so = frappe.db.sql(""" update `tabSales Order` set cluster = '%s' where name = '%s' """%(pp[6],pp[1]))
			print(pp[1])
		#     c = frappe.get_value("Sales Order",{"name":pp[2]},["docstatus"])
		#     if c == 0 or c == 1:
		#         print(pp[2])
	   
		# #         doc = frappe.get_doc("Sales Order",pp[2])
		# #         doc.prepared_by =  pp[6]
		# #         doc.sale_person =  pp[7]
		# #         doc.territory =  pp[8]
		# #         doc.save(ignore_permissions=True)
		#         so = frappe.db.sql(""" update `tabSales Order` set cluster = '%s' where name = '%s' """%(pp[6],pp[2]))
		#         print(so)

	   
@frappe.whitelist()
def add_specification(template):
	# frappe.errprint(template)
	it = frappe.get_doc("Inspection Template",template)
	# frappe.errprint(it.functional_aspects_table)
	return it.functional_aspects_table,it.visual_aspects_table,it.material_aspects_table,it.dimensional_aspects_table
	# frappe.errprint(it.dimensional_aspects_table)
	# for i in it.dimensional_aspects_table:
	#     frappe.errprint()
	# frappe.errprint(it.material_aspects_table)
	# frappe.errprint(it.functional_aspects_table)

# @frappe.whitelist()
# def get_sub_heading(item_detail):
#     item_detail = json.loads(item_detail)
#     l1 = []
#     l2 = []
#     l3 = []
#     l4 = []
#     l5 = []
#     l6 = []
#     for i in item_detail:
#         if i['item_heading'] not in l1:
#             l1.append(i["item_heading"])
#         else:
#             l2.append(i["item_heading"])
#         if i['item_sub_heading'] not in l3:
#             l3.append(i['item_sub_heading'])
#         else:
#             l4.append(i['item_sub_heading'])
#         if i['item_title_3'] not in l5:
#             l5.append(i['item_title_3'])
#         else:
#             l6.append(['item_title_3'])
#     return l1,l3,l5

@frappe.whitelist()
def update_company_currency_in_payments(selected_currency,account_currency):
	return get_exchange_rate(selected_currency,account_currency)


@frappe.whitelist()
def get_dimensional(doc):
	sample = frappe.get_all("Inspection Sample",{"item_inspection":doc.name},["*"])
	item = frappe.get_value("Item",{"name":doc.item_code},["inspection_template"])
	it = frappe.db.sql(""" select `tabDimensional Aspects Table`.specification from `tabInspection Template` 
	left join `tabDimensional Aspects Table` on `tabInspection Template`.name = `tabDimensional Aspects Table`.parent
	where  `tabInspection Template`.name = '%s' """ %(item),as_dict = True)
	data = ''
	data += '<tr><td colspan="18" style="border-color:#e20026"><center><b>Dimensional Aspects<b></center></td></tr><tr>'
	data += '<tr><td width = "20%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;padding-bottom :50px;" >,<center><b style = "color:white;" >Specification<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Acceptance  Criteria<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">Instrument/Ref No<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">1<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >2<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >3<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >4<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >5<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Remarks<b></center></td></tr><tr>'
	for s in it:
		data += '<tr><td style="border-color:#e20026">%s</td><td style="border-color:#e20026">''</td><td style="border-color:#e20026">''</td>'%(s.specification)
		for i in sample:
			spec = frappe.db.sql(""" select `tabDimensional Aspects Sample`.specification,`tabDimensional Aspects Sample`.acceptance_criteria ,`tabDimensional Aspects Sample`.status from `tabInspection Sample` 
			left join `tabDimensional Aspects Sample` on `tabInspection Sample`.name = `tabDimensional Aspects Sample`.parent
			where  `tabInspection Sample`.item_inspection = '%s' and `tabDimensional Aspects Sample`.specification = '%s' and `tabInspection Sample`.name = '%s' """ %(doc.name,s.specification,i.name),as_dict = True)
			frappe.errprint(spec[0])
			if spec[0]['status']:
				data +='<td colspan="1" style="border-color:#e20026">%s</td>'%('&#10004')
			else:
				data +='<td colspan="1" style="border-color:#e20026"><center>%s</center></td>'%('-')
		data += '<td style="border-color:#e20026">''</td>'
		data +='</tr>'
	return data

def get_functional(doc):
	sample = frappe.get_all("Inspection Sample",{"item_inspection":doc.name},["*"])
	item = frappe.get_value("Item",{"name":doc.item_code},["inspection_template"])
	it = frappe.db.sql(""" select `tabFunctional Aspects Table`.specification from `tabInspection Template` 
	left join `tabFunctional Aspects Table` on `tabInspection Template`.name = `tabFunctional Aspects Table`.parent
	where  `tabInspection Template`.name = '%s' """ %(item),as_dict = True)
	data = ''
	data += '<tr><td colspan="18" style="border-color:#e20026"><center><b>Functional Aspects<b></center></td></tr><tr>'
	data += '<tr><td width = "20%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;padding-bottom :50px;" >,<center><b style = "color:white;" >Specification<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Acceptance  Criteria<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">Instrument/Ref No<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">1<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >2<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >3<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >4<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >5<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Remarks<b></center></td></tr><tr>'
	for s in it:
		data += '<tr><td style="border-color:#e20026">%s</td><td style="border-color:#e20026">''</td><td style="border-color:#e20026">''</td>'%(s.specification)
		for i in sample:
			spec = frappe.db.sql(""" select `tabFunctional Aspects Sample`.specification,`tabFunctional Aspects Sample`.acceptance_criteria ,`tabFunctional Aspects Sample`.status from `tabInspection Sample` 
			left join `tabFunctional Aspects Sample` on `tabInspection Sample`.name = `tabFunctional Aspects Sample`.parent
			where  `tabInspection Sample`.item_inspection = '%s' and `tabFunctional Aspects Sample`.specification = '%s' and `tabInspection Sample`.name = '%s' """ %(doc.name,s.specification,i.name),as_dict = True)
			if spec[0]['status']:
				data +='<td colspan="1" style="border-color:#e20026">%s</td>'%('&#10004')
			else:
				data +='<td colspan="1" style="border-color:#e20026"><center>%s</center></td>'%('-')
		data += '<td style="border-color:#e20026">''</td>'
		data +='</tr>'
	return data

def get_material(doc):
	sample = frappe.get_all("Inspection Sample",{"item_inspection":doc.name},["*"])
	item = frappe.get_value("Item",{"name":doc.item_code},["inspection_template"])
	it = frappe.db.sql(""" select `tabMaterial Aspects Table`.specification from `tabInspection Template` 
	left join `tabMaterial Aspects Table` on `tabInspection Template`.name = `tabMaterial Aspects Table`.parent
	where  `tabInspection Template`.name = '%s' """ %(item),as_dict = True)
	data = ''
	data += '<tr><td colspan="18" style="border-color:#e20026"><center><b>Material Aspects<b></center></td></tr><tr>'
	data += '<tr><td width = "20%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;padding-bottom :50px;" >,<center><b style = "color:white;" >Specification<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Acceptance  Criteria<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">Instrument/Ref No<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">1<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >2<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >3<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >4<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >5<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Remarks<b></center></td></tr><tr>'
	for s in it:
		data += '<tr><td style="border-color:#e20026">%s</td><td style="border-color:#e20026">''</td><td style="border-color:#e20026">''</td>'%(s.specification)
		for i in sample:
			spec = frappe.db.sql(""" select `tabMaterial Aspects Sample`.specification,`tabMaterial Aspects Sample`.acceptance_criteria ,`tabMaterial Aspects Sample`.status from `tabInspection Sample` 
			left join `tabMaterial Aspects Sample` on `tabInspection Sample`.name = `tabMaterial Aspects Sample`.parent
			where  `tabInspection Sample`.item_inspection = '%s' and `tabMaterial Aspects Sample`.specification = '%s' and `tabInspection Sample`.name = '%s' """ %(doc.name,s.specification,i.name),as_dict = True)
			if spec[0]['status']:
				data +='<td colspan="1" style="border-color:#e20026">%s</td>'%('&#10004')
			else:
				data +='<td colspan="1" style="border-color:#e20026"><center>%s</center></td>'%('-')
		data += '<td style="border-color:#e20026">''</td>'
		data +='</tr>'
	return data

def get_visual(doc):
	# s = frappe.get_all("Inspection Sample",{"item_inspection":doc.name},["*"])
	item = frappe.get_value("Item",{"name":doc.item_code},["inspection_template"])
	it = frappe.db.sql(""" select `tabVisual Aspects Table`.specification from `tabInspection Template` 
	left join `tabVisual Aspects Table` on `tabInspection Template`.name = `tabVisual Aspects Table`.parent
	where  `tabInspection Template`.name = '%s' """ %(item),as_dict = True)
	data = ''
	data += '<tr><td colspan="10"><center><b>Visual Aspects</b></center></td></tr><tr>'
	data += '<tr><td width = "20%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;padding-bottom :50px;" >,<center><b style = "color:white;" >Specification<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Acceptance  Criteria<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">Instrument/Ref No<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white">1<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >2<b></center></td><td  width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >3<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >4<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;"><center><b style = "color:white" >5<b></center></td><td width = "10%" style = "border-color:#e20026;border-right-color:#e20026;background-color: #e20026;color: white;" ><center><b style = "color:white" >Remarks<b></center></td></tr><tr>'
	for s in it:
		data += '<tr><td style="border-color:#e20026">%s</td><td style="border-color:#e20026">''</td><td style="border-color:#e20026">''</td>'%(s.specification)
	return data

def india_quotation(doc):
	data = ''
	data += '<tr style="height:10%;">'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">No</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;width:18%;">Item</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;width:40%,height:10px;">Description</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">UOM</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Qty</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Rate </td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Amount</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Datasheet Link</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Remarks</td> </tr>'
	for h in doc.heading: 
		data += '<tr style="height:10%;">'
		data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b><center>%s</center></b></span></td></tr>' %(h.main_title)
		for s in doc.sub_heading:
			if h.main_title == s.heading:
				if s.title:
					data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b>%s</b></span></td></tr>' %(s.title)
				else:
					pass
				for t in doc.title_3:
					if h.main_title == t.heading and s.title == t.sub_heading:
						if t.title_3:
							data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b>%s</b></span></td></tr>' %(t.title_3)
						else:
							pass
						x = 1
						for i in doc.items:
							if i.item_heading == h.main_title and i.item_sub_heading == s.title and i.item_title_3 == t.title_3:
								if i.optional_item == 1:
									data+='<tr class ="altercolor" style="font-size:10px; height:10%;">'

									data+='<td style="font-family:Arial;border-color:#e20026;">%s</td>' %(x)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.item_code)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.description)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.uom)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.qty)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.rate)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.amount)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %('')
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td></tr>' %(i.remark or '')
								else:
									data+='<tr class ="altercolor" style="font-size:10px; height:10%;">'

									data+='<td style="font-family:Arial;border-color:#e20026;">%s</td>' %(x)
									data+='<td style="border-color:#e20026">%s</td>' %(i.item_code)
									data+='<td style="border-color:#e20026">%s</td>' %(i.description)
									data+='<td style="border-color:#e20026">%s</td>' %(i.uom)
									data+='<td style="border-color:#e20026">%s</td>' %(i.qty)
									data+='<td style="border-color:#e20026">%s</td>' %(i.rate)
									data+='<td style="border-color:#e20026">%s</td>' %(i.amount)
									data+='<td style="border-color:#e20026;">%s</td>' %('')
									data+='<td style="border-color:#e20026">%s</td></tr>' %(i.remark or '')

							x = x + 1
	return data


def opportunity_india(doc):
	data = ''
	data += '<tr style="height:10%;">'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">No</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;width:18%;">Item</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;width:40%,height:10px;">Description</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">UOM</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Qty</td>'
	# data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Rate </td>'
	# data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Amount</td>'
	# data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Datasheet Link</td>'
	data += '<td style="border-color:#e20026;border-right-color:white;background-color: #e20026;color: white;">Remarks</td> </tr>'
	for h in doc.heading: 
		data += '<tr style="height:10%;">'
		data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b><center>%s</center></b></span></td></tr>' %(h.main_title)
		for s in doc.sub_heading:
			if h.main_title == s.heading:
				if s.title:
					data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b>%s</b></span></td></tr>' %(s.title)
				else:
					pass
				for t in doc.title_3:
					if h.main_title == t.heading and s.title == t.sub_heading:
						if t.title_3:
							data += '<td style="border-color:#e20026;text-align:left" colspan="10"><span><b>%s</b></span></td></tr>' %(t.title_3)
						else:
							pass
						x = 1
						for i in doc.items:
							if i.item_heading == h.main_title and i.item_sub_heading == s.title and i.item_title_3 == t.title_3:
								if i.optional == 1:
									data+='<tr class ="altercolor" style="font-size:10px; height:10%;">'
									data+='<td style="font-family:Arial;border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(x)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.item_code)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.description)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.uom)
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td>' %(i.qty)
									# data+='<td style="border-color:#e20026">%s</td>' %(i.rate)
									# data+='<td style="border-color:#e20026">%s</td>' %(i.amount)
									# data+='<td style="border-color:#e20026;">%s</td>' %('')
									data+='<td style="border-color:#e20026;background-color: #F1EAE8;">%s</td></tr>' %(i.remarks or '')
								else:
									data+='<tr class ="altercolor" style="font-size:10px; height:10%;">'
									data+='<td style="font-family:Arial;border-color:#e20026;">%s</td>' %(x)
									data+='<td style="border-color:#e20026;">%s</td>' %(i.item_code)
									data+='<td style="border-color:#e20026">%s</td>' %(i.description)
									data+='<td style="border-color:#e20026">%s</td>' %(i.uom)
									data+='<td style="border-color:#e20026">%s</td>' %(i.qty)
									# data+='<td style="border-color:#e20026">%s</td>' %(i.rate)
									# data+='<td style="border-color:#e20026">%s</td>' %(i.amount)
									# data+='<td style="border-color:#e20026;">%s</td>' %('')
									data+='<td style="border-color:#e20026">%s</td></tr>' %(i.remarks or '')
							x = x + 1
	return data

@frappe.whitelist()
def quotation_workflow_alert(doc,method):
	wa = frappe.get_doc("Workflow Approval","Quotation")
	for i in wa.workflow:
		if doc.cluster:
			if doc.so_check == 0 and doc.cluster == i.cluster and doc.company == i.company and doc.price_list_region == i.territory and i.role == "Sales Manager" and doc.work_flow == "Pending for Sales Manager":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Quotation - %s is Pending for Approval' %(doc.name),
					message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)  
				doc.so_check = 1
			
			if doc.op_check == 0 and doc.cluster == i.cluster and doc.company == i.company and doc.price_list_region == i.territory and i.role == "Operation Director" and doc.work_flow == "Pending for Operation Director":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Quotation - %s is Pending for Approval' %(doc.name),
					message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   
				doc.op_check = 1

			if doc.coo_check == 0 and doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Quotation - %s is Pending for Approval' %(doc.name),
					message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)
				doc.coo_check =1
		
		else:
			if doc.so_check == 0 and doc.company == i.company and doc.price_list_region == i.territory and i.role == "Sales Manager" and doc.work_flow == "Pending for Sales Manager":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Quotation - %s is Pending for Approval' %(doc.name),
					message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   
				doc.so_check = 1
			
				if doc.op_check == 0 and doc.company == i.company and doc.price_list_region == i.territory and i.role == "Operation Director" and doc.work_flow == "Pending for Operation Director":
					frappe.sendmail(
						recipients=[i.user ],
						subject = 'Quotation - %s is Pending for Approval' %(doc.name),
						message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

					)
				doc.op_check = 1

				if doc.coo_check == 0 and doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
					frappe.sendmail(
						recipients=[i.user ],
						subject = 'Quotation - %s is Pending for Approval' %(doc.name),
						message = """ Quotation - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

					)
				doc.coo_check =1

@frappe.whitelist()
def so_workflow_alert(doc,method):
	wa = frappe.get_doc("Workflow Approval","Sales Order")
	for i in wa.workflow:
		if doc.cluster:
			if doc.cluster == i.cluster and doc.company == i.company and i.role == "HOD" and doc.workflow_state == "Pending for HOD":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
					message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   
			
			if doc.cluster == i.cluster and doc.company == i.company and i.role == "Accounts User" and doc.work_flow == "Pending for Accounts":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
					message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   

			# if doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
			#     frappe.sendmail(
			#         recipients=[i.user ],
			#         subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
			#         message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

			#     )
		
		else:
			if doc.company == i.company and i.role == "HOD" and doc.work_flow == "Pending for HOD":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
					message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   

			if doc.company == i.company and i.role == "Accounts User" and doc.work_flow == "Pending for Accounts":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Sales order - %s is Pending for Approval' %(doc.name),
					message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)

				# if doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
				#     frappe.sendmail(
				#         recipients=[i.user ],
				#         subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
				#         message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				#     )

@frappe.whitelist()
def po_workflow_alert(doc,method):
	wa = frappe.get_doc("Workflow Approval","Purchase Order")
	for i in wa.workflow:
		if doc.cluster:
			if doc.cluster == i.cluster and doc.company == i.company and i.role == "Purchase Master Manager" and doc.workflow_state == "Pending for purchase manager":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Purchase Order - %s is Pending for Approval' %(doc.name),
					message = """ Purchase Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/purchase-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   
			
			if doc.cluster == i.cluster and doc.company == i.company and i.role == "Accounts User" and doc.work_flow == "Pending for Accounts":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Purchase Order - %s is Pending for Approval' %(doc.name),
					message = """ Purchase Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/purchase-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   

			# if doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
			#     frappe.sendmail(
			#         recipients=[i.user ],
			#         subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
			#         message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/sales-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

			#     )
		
		else:
			if doc.company == i.company and i.role == "Purchase Master Manager" and doc.work_flow == "Pending for purchase manager":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Purchase Order - %s is Pending for Approval' %(doc.name),
					message = """ Purchase Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/purchase-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)   

			if doc.company == i.company and i.role == "Accounts User" and doc.work_flow == "Pending for Accounts":
				frappe.sendmail(
					recipients=[i.user ],
					subject = 'Purchase Order - %s is Pending for Approval' %(doc.name),
					message = """ Purchase Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/purchase-order/%s">Click here</a> to approve """ %(doc.name,doc.name)

				)

				# if doc.company == i.company and i.role == "COO" and doc.work_flow == "Pending for COO":
				#     frappe.sendmail(
				#         recipients=[i.user ],
				#         subject = 'Sales Order - %s is Pending for Approval' %(doc.name),
				#         message = """ Sales Order - %s is pending,waiting for your approval <a href = "https://erp.nordencommunication.com/app/quotation/%s">Click here</a> to approve """ %(doc.name,doc.name)

				#     )



@frappe.whitelist()
def check_item_inspection_dn(doc,method):
	for i in doc.items:
		if not i.skip_qc:
			item = frappe.get_value("Item",{"name":i.item_code},["request_for_quality_inspection"])
			inspection = frappe.db.exists("Item Inspection",{"item_code":i.item_code,"pr_mumber":doc.name,'docstatus':'1'})
			if item == 1:
				if not inspection:
					frappe.throw("Please inspect the items")
			else:
				frappe.validated = True

@frappe.whitelist()
def leave_notification():
	employee = frappe.db.get_all('Employee',{'status':'Active','Company':'Norden Communication Pvt Ltd'},['*'])
	for emp in employee:
		if emp.company_email:
			frappe.sendmail(
				recipients = emp.company_email,
				# 'api@groupteampro.com','hrd@nordencommunication.com','aiswarya@nordencommunication.com','kareem@nordencommunication.com'],
				subject = 'ERP Notification-Leave Application & Work from home',
				message = """<p style='font-size:15px'>Dear All,</p>
						<p align:'justify'>On behalf of NORDEN Management, HRD is very pleased to introduce you to an online <b>Leave Application & Work from Home</b> process through our new ERP system.</p>
						<p align:'justify'>Effective 06th January 2023, all <b>Leave Application & Work from home requests</b> can be processed online (submission and approval). However you will be able to monitor the status of your leave balance & request.</p>
						<p align:'justify'>HR & Admin Expert- Ms.Aiswarya will be able to clarify the process in case you need any support. Meanwhile please check your leave balance in ERP in the beginning itself , if you have any queries on it, please coordinate with Ms. Asha immediately</p>
						<p align:'justify'>NB:as informed before, Hereafter there is no carry forward option available for casual & sick leave, both are fixed as 12 days for all permanent staff.</p>
						<table border="1">
						<tr><td>ERP WEBSITE LINK</td><td><a href="https://erp.nordencommunication.com">erp.nordencommunication.com</a></td></tr>
						<tr><td><b>Leave Application Process Link</b></td><td><a href="https://www.awesomescreenshot.com/video/13751301?key=cf795cbc77b91c6b93a571e5b45fd477">Leave Application</a></td></tr>
						<tr><td><b>Work from home and leave application Process Link</b></td><td><a href="https://www.awesomescreenshot.com/video/13753132?key=b67e492b31bafea03e78c30a4c2f548a">Work From Home and Attendance Request</a></td></tr>
						</table><br>
						<p>Thanking You</p>
						<p>HRD</p>
						"""
			)
		else:
			message = ('Company Email Not in Employee %s'%(emp.name))   
			frappe.log_error('Email Notification',message) 
		
@frappe.whitelist()
def pr_creation(doc,method):
	if doc.logistics:
		lr = frappe.get_doc("Logistics Request",doc.logistics)
		for i in doc.receipts:
			i.purchase_receipt = lr.name
		

@frappe.whitelist()
def set_cluster(sales_person):
	doc = frappe.get_value("Cluster",{"user":sales_person})
	if doc:
		return doc
   

@frappe.whitelist()
def opp_details(opp):
	doc = frappe.get_doc("Opportunity",opp)
	return doc.items
	# for i in doc.heading:
	#     frappe.errprint(i.main_title)
	
@frappe.whitelist()
def create_sti(doc,method):
	if doc.bill_of_entry or doc.bom_out_bound:
		if doc.stock_entry_type == "Material Receipt":
			if frappe.db.exists("Stock Transfer India",doc.bill_of_entry):
				s = frappe.get_doc("Stock Transfer India",doc.bill_of_entry)
				s.boe_inbound_no = doc.bill_of_entry
				for i in doc.items:
					x = s.append("stock_transfer")
					x.item_code = i.item_code
					x.target_warehouse = i.t_warehouse
					x.bom_inbound_qty = i.qty
					x.balance_in_qty = i.qty
					x.bom_inbound = doc.bill_of_entry
					x.serial_no = i.serial_no
					x.qty = i.qty
					x.uom = i.uom
					x.basic_rate = i.basic_rate
					x.additional_cost = i.additional_cost
					x.itemwise_additional_cost = i.itemwise_additional_cost
					x.valuation_rate = i.valuation_rate
					x.basic_amount = i.basic_amount
					x.amount = i.amount
					x.batch = i.batch_no
				s.save(ignore_permissions=True)
			else:
				s = frappe.new_doc("Stock Transfer India")
				s.warehouse = doc.to_warehouse
				s.boe_inbound_no = doc.bill_of_entry
				for i in doc.items:
					x = s.append("stock_transfer")
					x.item_code = i.item_code
					x.target_warehouse = i.t_warehouse
					x.bom_inbound_qty = i.qty
					x.balance_in_qty = i.qty
					x.bom_inbound = doc.bill_of_entry
					x.serial_no = i.serial_no
					x.updated_serial_no = i.serial_no
					x.qty = i.qty
					x.uom = i.uom
					x.basic_rate = i.basic_rate
					x.additional_cost = i.additional_cost
					x.itemwise_additional_cost = i.itemwise_additional_cost
					x.valuation_rate = i.valuation_rate
					x.basic_amount = i.basic_amount
					x.amount = i.amount
					x.batch = i.batch_no
				s.save(ignore_permissions=True)

		if doc.stock_entry_type == "Material Transfer" and doc.bom_out_bound:
			s = frappe.new_doc("Stock Transfer India")
			s.boe_outbound_no = doc.bom_out_bound
			for i in doc.items:
				x = s.append("stock_transfer")
				x.item_code = i.item_code
				x.source_warehouse = i.s_warehouse
				x.target_warehouse = i.t_warehouse
				x.bom_outbound_qty = i.qty
				x.balance_out_qty = i.qty
				x.bom_outbound = doc.bom_out_bound
				x.serial_no = i.serial_no
				x.qty = i.qty
				x.uom = i.uom
				x.basic_rate = i.basic_rate
				x.additional_cost = i.additional_cost
				x.itemwise_additional_cost = i.itemwise_additional_cost
				x.valuation_rate = i.valuation_rate
				x.basic_amount = i.basic_amount
				x.amount = i.amount
				x.batch = i.batch_no
			s.save(ignore_permissions=True)
			s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
			for i in s.stock_transfer:
				for d in doc.items:
					if d.item_code == i.item_code and i.target_warehouse == "India Bonded Warehouse - NCPL":
						i.balance_in_qty = i.balance_in_qty - d.qty 
						# out_list = json.loads(doc.bom_out_bound)
						# i.outbound_list = out_list
						if i.updated_serial_no:
							s_name = (i.updated_serial_no).upper() 
							ser_name = s_name.split("\n")
							d_name = (d.serial_no).upper() 
							sr_name = d_name.split("\n")
							for k in sr_name:
								if k in ser_name:
									ser_name.remove(k)
							i.updated_serial_no = ''
							i.updated_serial_no = "\n".join(ser_name)
			s.save(ignore_permissions=True)

@frappe.whitelist()
def reversing_se(doc,method):
    if doc.bill_of_entry and doc.stock_entry_type == "Material Transfer":
        s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
        for i in s.stock_transfer:
            for d in doc.items:
                if d.item_code == i.item_code and i.target_warehouse == "India Bonded Warehouse - NCPL":
                    i.balance_in_qty = i.balance_in_qty + d.qty 
                    if i.updated_serial_no:
                        s_name = (i.updated_serial_no).upper() 
                        ser_name = s_name.split("\n")
                        d_name = (d.serial_no).upper()
                        sr_name = d_name.split("\n")
                        for k in sr_name:
                            ser_name.append(k)
                            frappe.errprint("\n".join(ser_name))
                        i.updated_serial_no = ''
                        i.updated_serial_no = "\n".join(ser_name)
        s.save(ignore_permissions=True)
        o = frappe.get_doc("Stock Transfer India",{"boe_outbound_no":doc.bom_out_bound})
        o.save(ignore_permissions=True)
        o.delete()
        

    if doc.bill_of_entry and doc.stock_entry_type == "Material Receipt":
        s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
        s.save(ignore_permissions=True)
        s.delete()

def add_itemwise_additional_cost(doc,method):
	for row in doc.items:
		if row.itemwise_additional_cost:
			row.additional_cost += row.itemwise_additional_cost

@frappe.whitelist()
def update_sn(doc,method):
	for i in doc.items:
		if i.serial_no:
			s_name = (i.serial_no).upper() 
			ser_name = s_name.split("\n")
			for sn in ser_name:
				if frappe.db.exists("Serial No",sn):
					frappe.db.set_value("Serial No",sn,"inbound",doc.bill_of_entry)
					frappe.db.set_value("Serial No",sn,"outbound",doc.bom_out_bound)          

@frappe.whitelist()
def update_sn_pr(doc,method):
	for i in doc.items:
		if i.serial_no:
			s_name = (i.serial_no).upper() 
			ser_name = s_name.split("\n")
			for sn in ser_name:
				frappe.errprint(sn)
				if frappe.db.exists("Serial No",sn):
					frappe.db.set_value("Serial No",sn,"inbound",doc.bill_of_entry)
					# frappe.db.set_value("Serial No",sn,"outbound",doc.bom_out_bound)               


@frappe.whitelist()
def currency_conversion_lg(gt,cr):
	ep = get_exchange_rate(cr,"INR")
	s = float(gt) * float(ep)
	return s

@frappe.whitelist()
def check_uom(item,uom):
	s = frappe.get_doc("Item",item)
	return s.uoms
   
@frappe.whitelist()
def set_inbound(doc,method):
	if doc.bill_of_entry and doc.bom_out_bound:
		if frappe.db.exists("Bill of Entry Outbound",doc.bom_out_bound):
			s = frappe.get_doc("Bill of Entry Outbound",doc.bom_out_bound)
			s.boe_inbound = doc.bill_of_entry
			s.save(ignore_permissions=True)

@frappe.whitelist()
def return_blocked_items():
	block = frappe.get_all("Block Warehouse",{"workflow_state":"Transfered"},["*"])
	for i in block:
		days =  add_days(days,15)
		if days > today():
			print(i.name)
			b = frappe.get_doc("Block Warehouse",i.name)
			b.workflow_state = "Returned"
			b.status = "Return"
			stock = frappe.new_doc("Stock Entry")
			stock.company = b.company
			stock.stock_entry_type = "Material Transfer"
			stock.from_warehouse = b.target
			stock.to_warehouse = b.source
			for s in b.items:
				stock.append("items", {
				"s_warehouse":b.target,
				"t_warehouse": b.source,
				"item_code": s.item_code,
				"qty":s.quantity,
				"allow_zero_valuation_rate":1
				})
			stock.save(ignore_permissions=True)
			stock.submit()

@frappe.whitelist()
def create_stock_transfer_india(doc,method):
	if doc.bill_of_entry and doc.set_warehouse == "India Bonded Warehouse - NCPL":
		if frappe.db.exists("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry}):
			s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
			for i in doc.items:
				x = s.append("stock_transfer")
				x.item_code = i.item_code
				x.item_name = i.item_name
				x.target_warehouse = i.warehouse
				x.bom_inbound_qty = i.received_qty
				x.balance_in_qty = i.received_qty
				x.bom_inbound = doc.bill_of_entry
				x.serial_no = i.serial_no
				x.updated_serial_no = i.serial_no
				x.qty = i.received_qty
				x.uom = i.uom
				x.basic_rate = i.rate
				# x.additional_cost = i.additional_cost
				# x.itemwise_additional_cost = i.itemwise_additional_cost
				# x.valuation_rate = i.valuation_rate
				x.basic_amount = i.amount
				x.amount = i.amount
				x.batch = i.batch_no
			s.save(ignore_permissions=True)
		else:
			s = frappe.new_doc("Stock Transfer India")
			s.boe_inbound_no = doc.bill_of_entry
			for i in doc.items:
				x = s.append("stock_transfer")
				x.item_code = i.item_code
				x.item_name = i.item_name
				x.target_warehouse = i.warehouse
				x.bom_inbound_qty = i.received_qty
				x.balance_in_qty = i.received_qty
				x.bom_inbound = doc.bill_of_entry
				x.serial_no = i.serial_no
				x.updated_serial_no = i.serial_no
				x.qty = i.received_qty
				x.uom = i.uom
				x.basic_rate = i.rate
				# x.additional_cost = i.additional_cost
				# x.itemwise_additional_cost = i.itemwise_additional_cost
				# x.valuation_rate = i.valuation_rate
				x.basic_amount = i.amount
				x.amount = i.amount
				x.batch = i.batch_no
			s.save(ignore_permissions=True)

	if doc.bill_of_entry and doc.boe_outbound:
		s = frappe.new_doc("Stock Transfer India")
		s.boe_outbound_no = doc.boe_outbound
		for i in doc.items:
			x = s.append("stock_transfer")
			x.item_code = i.item_code
			x.item_name = i.item_name
			x.target_warehouse = i.warehouse
			x.bom_inbound_qty = i.received_qty
			x.balance_in_qty = i.received_qty
			x.bom_inbound = doc.bill_of_entry
			x.serial_no = i.serial_no
			x.updated_serial_no = i.serial_no
			x.qty = i.received_qty
			x.uom = i.uom
			x.basic_rate = i.rate
			# x.additional_cost = i.additional_cost
			# x.itemwise_additional_cost = i.itemwise_additional_cost
			# x.valuation_rate = i.valuation_rate
			x.basic_amount = i.amount
			x.amount = i.amount
			x.batch = i.batch_no
		s.save(ignore_permissions=True)
		s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
		for i in s.stock_transfer:
			ob = []
			out = (i.outbound_list).upper()
			out_name = out.split("\n")
			ob.extend(out_name)
			frappe.errprint(ob)
			for d in doc.items:
				if d.item_code == i.item_code and i.target_warehouse == "India Bonded Warehouse - NCPL":
					i.balance_in_qty = i.balance_in_qty - d.qty 
					ob.append(doc.boe_outbound)
					out_list = ob
					i.outbound_list = "\n".join(out_list)
					s_name = (i.updated_serial_no).upper() 
					ser_name = s_name.split("\n")
					frappe.errprint(ser_name)
					d_name = (d.serial_no).upper() 
					sr_name = d_name.split("\n")
					for k in sr_name:
						if k in ser_name:
							ser_name.remove(k)
					i.updated_serial_no = ''
					i.updated_serial_no = "\n".join(ser_name)
		s.save(ignore_permissions=True)

@frappe.whitelist()
def reverse_sti_pr(doc,method):
    if doc.bill_of_entry:
        s = frappe.get_doc("Stock Transfer India",{"boe_inbound_no":doc.bill_of_entry})
        s.save(ignore_permissions=True)
        s.delete()

@frappe.whitelist()
def transfer_to_scrap(doc,method):
	if doc.mrb_action == "Scrap":
		s = frappe.new_doc("Stock Entry")
		s.stock_entry_type = "Material Transfer"   
		s.from_warehouse = doc.warehouse     
		scrap_warehouse = frappe.get_value("Warehouse",{"company":doc.company,"is_scrap":1})
		s.to_warehouse = scrap_warehouse
		for i in s.items:
			i.item_code = doc.item_code
			i.qty = doc.rejected_qty
			i.uom = doc.uom
			i.batch_no = doc.batch_no
			i.basic_rate = doc.rate
		s.save(ignore_permissions=True)


@frappe.whitelist()
def po_file_summary(so,mr,supplier,currency,item_details):
	po = frappe.get_all("Purchase Order",{"sales_order_number":so},["*"])
	item_details = json.loads(item_details)
	cus = frappe.get_value("Sales Order",{"name":so},["customer"])
	gt = frappe.get_value("Sales Order",{"name":so},["grand_total"])
	pt = frappe.get_value("Sales Order",{"name":so},["Payment_terms_template"])
	icm = frappe.get_value("Sales Order",{"name":so},["internal_cost_margin"])
	outstanding = frappe.get_all("Sales Invoice",{"customer":cus},["*"])
	s_out = 0
	for i in outstanding:
		if not i.status == "Paid":
			s_out = s_out + i.outstanding_amount
	sq = frappe.db.sql(""" select `tabMaterial Request Item`.item_code as item_code,`tabMaterial Request Item`.sales_order_qty as sales_order_qty,`tabMaterial Request Item`.stock_qty as stock_qty 
	from `tabMaterial Request` left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent where `tabMaterial Request`.name = '%s' """ %(mr),as_dict=True)
	# so_total = 0
	# s_total = 0
	# for i in sq:
	#     for s in item_details:
	#         if i.item_code == s["item_code"]:
	#             so_total = so_total + i.sales_order_qty * s["rate"]
	#             s_total = s_total + i.stock_qty * s["rate"]
	data = ''
	data+='<table class="table"><style>table,tr,td { padding:5px;border: 1px solid black; font-size:11px;border-color:#e20026;} </style>'
	# data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;" colspan=20><center><b>File Summary<b></center></td></tr>'
	data+='<tr style="background-color:lightgrey" ><td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;" colspan=20><center><b>Customer Details</b></center></td></tr>'
   
	data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;font-weight:bold;"><center>Name</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>'%(cus)
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;font-weight:bold;"><center>Margin</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>'%(round(icm,2))
	data+='</tr>'

	data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;font-weight:bold;"><center>Amount</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(gt)
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;font-weight:bold;"><center>Outstanding</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(round(s_out,2))
	data+='</tr>'

	data+='<tr><td style="padding:1px;border: 1px   solid black;font-size:14px;width:25%;border-color:#e20026;font-weight:bold;"><center>Payment Terms</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(pt or '')
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;"><center></center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026;"><center></center></td>'
	data+='</tr>'

	data+='<tr style="background-color:lightgrey" ><td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;" colspan=20><center><b>Supplier Details - B2B</b></center></td></tr>'
	
	data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Supplier</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Currency</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Amount-B2B</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Payment Terms</center></td>'
	data+='</tr>'

	for i in po:
		pos = frappe.db.sql(""" select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.rate as rate
		from `tabPurchase Order` left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent where `tabPurchase Order`.name = '%s' """ %(i.name),as_dict=True)
		sales_total = 0
		for s in pos:
			for r in sq:
				if s.item_code == r.item_code:
					sales_total = sales_total + r.sales_order_qty * s.rate
		data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.supplier)
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.currency)
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(sales_total)
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.payment_terms_template or '')
		data+='</tr>'

	data+='<tr style="background-color:lightgrey" ><td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;" colspan=20><center><b>Supplier Details - Stock</b></center></td></tr>'
	
	data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Supplier</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Currency</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Amount-Stock</center></td>'
	data+='<td style="padding:1px;border: 1px solid black;font-size:14px;width:25%;border-color:#e20026; font-weight: bold;"><center>Payment Terms</center></td>'
	data+='</tr>'
	for i in po:
		pos = frappe.db.sql(""" select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.rate as rate
		from `tabPurchase Order` left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent where `tabPurchase Order`.name = '%s' """ %(i.name or ""),as_dict=True)
		
		frappe.errprint(i.name)
		stock_total = 0
		for r in sq:
			for s in pos:
				if s.item_code == r.item_code:
					stock_total = stock_total + r.stock_qty * s.rate
		
		data+='<tr><td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.supplier)
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.currency)
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(round(stock_total,2))
		data+='<td style="padding:1px;border: 1px solid black;font-size:14px;border-color:#e20026;"><center>%s</center></td>' %(i.payment_terms_template or '')
		data+='</tr>'
	# stock_total = 0

	data+='</table>'
	return data


@frappe.whitelist()
def item_allocation(item_details,name,company):
	item_details = json.loads(item_details)
	ia = frappe.new_doc("Item Allocation")
	ia.document = "Quotation"
	ia.quotation = name

	for i in item_details:
		sno = frappe.get_all("Serial No",{"item_code":i["item_code"]})
		ia.append('allocation_item',{
			'item_code':i["item_code"],
			'item_name':i["item_name"],
			'qty':i["qty"],
		
		})
			
	ia.save(ignore_permissions = True)
	ai = frappe.get_doc("Item Allocation",{"name":ia.name})
	for i in item_details:
		sno = frappe.get_all("Serial No",{"item_code":i["item_code"]})
		for s in sno:
			for d in ai.allocation_item:
				# s_name = (s.name).upper()
				# name_s = s_name.split("\n")
				frappe.errprint(s.name)
				
			# d.serial_no = "\n".join(name_s)
				
	ai.save(ignore_permissions = True)

	frappe.msgprint("Items Transfered")


	# default_warehouse = frappe.get_value("Warehouse",{"company":company,"default_for_stock_transfer":1})
	# target_warehouse = frappe.get_value("Warehouse",{"company":company,"is_block":1})
	# stock = frappe.new_doc("Stock Entry")
	# stock.company = company
	# stock.stock_entry_type = "Material Transfer"
	# stock.from_warehouse = default_warehouse
	# stock.to_warehouse = target_warehouse
	
	# for i in item_details:
		
	#     stock.append("items", {
	#     "s_warehouse": default_warehouse,
	#     "t_warehouse": target_warehouse,
	#     "item_code": i["item_code"],
	#     "qty":i["qty"],
	#     "allow_zero_valuation_rate":1
	#     })
	# stock.save(ignore_permissions=True)
	
	# stock.submit()

# @frappe.whitelist()
# def cancel_stock():
#     se = frappe.db.sql("""update `tabStock Entry` set docstatus = 1 where name = 'SE-NCMEF-2023-00001' """)

@frappe.whitelist()
def create_product_testing(doc,method):
	for i in doc.items:
		pt = frappe.new_doc("Product Testing")
		pt.purchase_receipt = i.name
		pt.item_code = i.item_code
		pt.item_name = i.item_name
		pt.qty = i.qty
		pt.serial_no = i.serial_no
		pt.batch_no = i.batch_no
		pt.warehouse = i.rejected_warehouse
		pt.cost_center = i.cost_center
		pt.save(ignore_permissions = True)

@frappe.whitelist()
def update_previous_leave_allocation_manually():
	employee = frappe.db.sql(""" select date_of_joining,name from `tabEmployee` where status = 'Active' and employee_number like '1-01-%' and employee_number != "1-01-100" and company != 'Norden Communication Middle East FZE' """,as_dict=True)
	current_date = date.today()
	for i in employee:
		rdelta = relativedelta.relativedelta(current_date, i.date_of_joining)
		months = rdelta.years * 12 + rdelta.months
		if months >= 13:
			earned = frappe.db.sql(""" select name from `tabLeave Allocation` where leave_type = 'Earned Leave' and year(from_date) = year(current_date) and employee ='%s'  """%(i.name),as_dict=True)
			for j in earned:
				earn = frappe.get_doc("Leave Allocation",j.name)
				earn.new_leaves_allocated = earn.new_leaves_allocated + 1
				earn.save(ignore_permissions=True)

			exist = frappe.db.exists("Leave Allocation",{"leave_type":"Earned Leave","employee":i.name,"docstatus":1})
			if not exist:
				new_all = frappe.new_doc("Leave Allocation")
				new_all.employee = i.name
				new_all.new_leaves_allocated = 1
				new_all.leave_type = "Earned Leave"
				new_all.company =frappe.db.get_value("Employee",i.name,["company"])
				new_all.from_date = date.today()
				new_all.to_date = date.today().replace(month=12, day=31)
				new_all.save(ignore_permissions=True)
				new_all.submit()
    
@frappe.whitelist()
def cron_job_allocation():
	job = frappe.db.exists('Scheduled Job Type', 'update_previous_leave_allocation_manually')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")  
		sjt.update({
			"method" : 'norden.utils.update_previous_leave_allocation_manually',
			"frequency" : 'Monthly',
			# "cron_format" : '30 00 * * *'
		})
		sjt.save(ignore_permissions=True)
  
from datetime import datetime
import calendar
@frappe.whitelist()
def create_update_leave_allocation():
	employees = frappe.get_all("Employee",{"status":"Active",'Company':'Norden Communication Middle East FZE'},["*"])
	current_date = datetime.now().date()
	for emp in employees:
		doj = emp.date_of_joining
		diff = current_date - doj
		years = diff.days / 365.25  
		frappe.errprint(emp.name)
		print(int(years))
		if(int(years)) > 0 :
			if frappe.db.exists("Leave Allocation",{'employee':emp.employee,'leave_type':"Annual Leave"}):
				la = frappe.get_doc("Leave Allocation",{'employee':emp.employee,'leave_type':"Annual Leave"},["*"])
				la.new_leaves_allocated = la.new_leaves_allocated + 2.5
				la.to_date = "2100-12-31"
				la.save(ignore_permissions=True)
				la.submit()   
			else:
				la = frappe.new_doc("Leave Allocation")
				la.employee = emp.name
				la.leave_type = "Annual Leave"
				la.new_leaves_allocated = 2.5
				la.from_date = current_date
				la.to_date = "2100-12-31"
				la.save(ignore_permissions=True)
				la.submit()  
    
@frappe.whitelist()
def cron_job_allocation_1():
	job = frappe.db.exists('Scheduled Job Type', 'create_update_leave_allocation')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")  
		sjt.update({
			"method" : 'norden.utils.create_update_leave_allocation',
			"frequency" : 'Monthly',
			# "cron_format" : '30 00 * * *'
		})
		sjt.save(ignore_permissions=True)
  

import frappe
from datetime import datetime, timedelta , date
@frappe.whitelist()
def annual_leave_expired():
	employees = frappe.get_all("Employee",{'status':"Active","company":"Norden Communication Middle East FZE"},['*'])
	for emp in employees:
		if frappe.db.exists("Leave Allocation",{'employee':emp.name,'leave_type':"Annual Leave","docstatus":1}):
			today = date.today()
			leave = frappe.get_doc("Leave Allocation",{'employee':emp.name,'leave_type':"Annual Leave","docstatus":1})
			if leave.expiry_date == today:
				application = frappe.db.sql("""select * from `tabLeave Application` where  leave_type = 'Annual leave' and docstatus = 1 and from_date = '%s' and to_date = '%s' """%(leave.leave_start_date,leave.expiry_date),as_dict=1)[0]
				total = 0
				for a in application:
					total = a.total_leave_days
				if leave.expire_from > total :
					el = leave.expire_from - total
					leave.new_leaves_allocated = leave.new_leaves_allocated - el
					leave_end_date = str(leave.leave_end_date)
					do = datetime.strptime(leave_end_date, '%Y-%m-%d')
					add_do = do + timedelta(days=1)
					add_do_str = add_do.strftime("%Y-%m-%d")
					add_date = datetime.strptime(add_do_str, '%Y-%m-%d').date()
					leave.leave_start_date = add_date
					leave_end_date = str(leave.leave_end_date)
					do = datetime.strptime(leave_end_date, '%Y-%m-%d')
					add_do = do + timedelta(months=12)
					add_do_str = add_do.strftime("%Y-%m-%d")
					add_date = datetime.strptime(add_do_str, '%Y-%m-%d').date()
					leave.leave_end_date = add_date
					leave_end_date = str(leave.leave_end_date)
					do = datetime.strptime(leave_end_date, '%Y-%m-%d')
					add_do = do + timedelta(year=1) + timedelta(months=6)
					add_do_str = add_do.strftime("%Y-%m-%d")
					add_date = datetime.strptime(add_do_str, '%Y-%m-%d').date()
					leave.expiry_date = add_date
					leave.expire_from = 30
					leave.save(ignore_permissions = True)
					frappe.db.commit


@frappe.whitelist()
def cron_job():
	job = frappe.db.exists('Scheduled Job Type', 'annual_leave_expired')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")  
		sjt.update({
			"method" : 'norden.utils.annual_leave_expired',
			"frequency" : 'Cron',
			"cron_format" : '30 00 * * *'
		})
		sjt.save(ignore_permissions=True)
  

@frappe.whitelist()  
def si_name(filename):
	from frappe.utils.file_manager import get_file
	_file = frappe.get_doc("File", {"file_name":filename})
	filepath = get_file(filename)
	ips = read_csv_content(filepath[1])
	for ip in ips:
		sales_invoice = ip[0]
		si_name = frappe.get_doc("Sales Invoice",sales_invoice)
		for s in si_name.items:
			sales = s.sales_order
			so_name = frappe.get_doc("Sales Order",sales)
			for si in so_name.items:
				quote = si.prevdoc_docname
		print(sales_invoice)
		print(sales)
		print(quote)
		if quote:
			quote_update = frappe.get_doc("Quotation",quote)
			for quo in quote_update.items:
				disc_amt = quo.discount_rate * quo.qty
			# 	# frappe.db.set_value("Sales Order Item", {"parent":i.parent,"item_code":k.item_code}, 'unit_price_document_currency', k.unit_price_document_currency)
			# 	# frappe.db.set_value("Sales Order Item", {"parent":i.parent,"item_code":k.item_code}, 'discount_value', k.discount_value)
				frappe.db.set_value("Quotation Item", {"parent":quote,"item_code":quo.item_code}, 'disc_amt', disc_amt)
			# 	# frappe.db.set_value("Sales Order Item", {"parent":i.parent,"item_code":k.item_code}, 'disc_amt', disc_amt)
		if sales:
			sale = frappe.get_doc("Sales Order",sales)
			for sa in sale.items:
				quot = frappe.get_doc("Quotation",sa.prevdoc_docname)
				for qu in quot.items:
					if sa.item_code == qu.item_code:
						frappe.db.set_value("Sales Order Item", {"parent":sales,"item_code":qu.item_code}, 'discount_value', qu.discount_value)
						frappe.db.set_value("Sales Order Item", {"parent":sales,"item_code":qu.item_code}, 'discount', qu.discount)
						frappe.db.set_value("Sales Order Item", {"parent":sales,"item_code":qu.item_code}, 'discount_rate', qu.discount_rate)
						frappe.db.set_value("Sales Order Item", {"parent":sales,"item_code":qu.item_code}, 'disc_amt', qu.disc_amt)
						if qu.unit_price_document_currency:
							frappe.db.set_value("Sales Order Item", {"parent":sales,"item_code":qu.item_code}, 'unit_price_document_currency', qu.unit_price_document_currency)
		if sales:
			sale_in = frappe.get_doc("Sales Invoice",sales_invoice)
			for sin in sale_in.items:
				so_ord = frappe.get_doc("Sales Order",sin.sales_order)
				for sd in so_ord.items:
					if sin.item_code == sd.item_code:
						if sd.unit_price_document_currency:
							frappe.db.set_value("Sales Invoice Item", {"parent":sales_invoice,"item_code":sd.item_code}, 'unit_price_document_currency', sd.unit_price_document_currency)
						frappe.db.set_value("Sales Invoice Item", {"parent":sales_invoice,"item_code":sd.item_code}, 'discount', sd.discount)
						frappe.db.set_value("Sales Invoice Item", {"parent":sales_invoice,"item_code":sd.item_code}, 'discount_rate', sd.discount_rate)
						frappe.db.set_value("Sales Invoice Item", {"parent":sales_invoice,"item_code":sd.item_code}, 'discount_value', sd.discount_value)
						frappe.db.set_value("Sales Invoice Item", {"parent":sales_invoice,"item_code":sd.item_code}, 'disc_amt', sd.disc_amt)

@frappe.whitelist()  
def item_price_bulk_upload_csv(filename):
	from frappe.utils.file_manager import get_file
	_file = frappe.get_doc("File", {"file_name":filename})
	filepath = get_file(filename)
	ips = read_csv_content(filepath[1])
	no_item = ["HI"]
	for ip in ips:
		if frappe.db.exists('Item Price',{'item_code':ip[0]}):
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Cost Rate - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Cost Rate - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[2]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Landing - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Landing - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[3]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Incentive - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Incentive - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[4]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Internal - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Internal - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[5]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Base Sales Price - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Base Sales Price - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[6]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Retail - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Retail - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[7]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Dist. Price - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Dist. Price - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[8]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Electra Qatar - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Electra Qatar - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[9]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Project Group - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Project Group - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[10]
			
			if frappe.db.exists('Item Price',{'item_code':ip[0],'price_list':"Saudi Dist. - NCMEF"}):
				rate = frappe.get_doc('Item Price',{'item_code':ip[0],'price_list':"Saudi Dist. - NCMEF"})
				print(rate.price_list_rate)
				rate.price_list_rate = ip[11]
		else:
			no_item.append(lis)
	frappe.log_error(title="No Item",message=no_item)

		
def cancel_po():
	doc = frappe.get_doc('Purchase Receipt',"PR-NCPLP-2023-00034")
	doc.flags.ignore_validate = True
	doc.cancel()
	frappe.db.commit()




# def create_hooks_mark_ot():
#     job = frappe.db.exists('Scheduled Job Type', 'check_se_and_send_mail_quote')
#     if not job:
#         sjt = frappe.new_doc("Scheduled Job Type")
#         sjt.update({
#             "method": 'norden.custom.check_se_and_send_mail_quote',
#             "frequency": 'Cron',
#             "cron_format": '0 2 * * *'
#         })
#         sjt.save(ignore_permissions=True)



@frappe.whitelist()
def get_detailed_quo(quotation):
	quo = frappe.get_doc("Quotation",quotation)
	return quo.items

@frappe.whitelist()
def get_detailed_so(sales_order):
	so = frappe.get_doc("Sales Order",sales_order)
	return so.items

#For downloading item table as excel
@frappe.whitelist()
def make_item_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	doc = frappe.get_doc("Quotation",args.name)
	if doc:
		ws.append(["Item Code","Item Name","Qty","Unit Rate","Margin Percentage","Margin Rate","Margin Value","Discount","Discount Rate","Discount Value","Discount Amount","Amount"])
		for i in doc.items:
			ws.append([i.item_code,i.item_name,i.qty,i.unit_price_document_currency,i.margin_percentage,i.margin_rate,i.margin_value,i.discount,i.discount_rate,i.discount_value,i.disc_amt,i.amount])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary' 	

@frappe.whitelist()
def get_salary_structure_assignment_base(employee,docstatus):
	assignment = frappe.get_value(
		"Salary Structure Assignment",
		filters={
			"employee": employee,
			"docstatus": 1  # Consider only submitted assignments
		},
		fieldname="base",
		order_by="creation DESC",
		
	)

	return assignment or 0.0

@frappe.whitelist(allow_guest=True)
def get_electra_item(item):
	url = "https://erp.electraqatar.com/api/method/electra.custom.get_norden_details?item=%s" % (item)
	headers = { 'Content-Type': 'application/json','Authorization': 'token 583681db58de9d7:c9fc006b5b31cef'}
	# params = {"limit_start": 0,"limit_page_length": 20000}

	response = requests.request('GET',url,headers=headers)
	res = json.loads(response.text)
	return res

@frappe.whitelist(allow_guest=True)
def electra_item(item):
	url = "https://erp.electraqatar.com/api/method/electra.custom.norden_details?item=%s" % (item)
	headers = { 'Content-Type': 'application/json','Authorization': 'token 583681db58de9d7:c9fc006b5b31cef'}
	# params = {"limit_start": 0,"limit_page_length": 20000}

	response = requests.request('GET',url,headers=headers)
	res = json.loads(response.text)
	return res


@frappe.whitelist()
def get_basic_details(item_code,company):
	data = ''
	data+= '<br><table width = "100%"><style>td { text-align:left } table,tr,td { padding:5px;border: 1px solid black; font-size:11px;} </style>'

	data+=	'<tr><td ><b>ITEM CODE</b></td>'
	data += '<td align = "center" ><b>%s</b></td>' %(item_code)
	
	stocks = frappe.db.sql("""select sum(`tabBin`.actual_qty) as actual_qty from `tabBin`
							join `tabWarehouse` on `tabWarehouse`.name = `tabBin`.warehouse
							join `tabCompany` on `tabCompany`.name = `tabWarehouse`.company
							where `tabBin`.item_code = '%s' and `tabWarehouse`.company = '%s' """ % (item_code,company), as_dict=True)[0]
 
	if not stocks['actual_qty']:
			stocks['actual_qty'] = 0
   
	if stocks["actual_qty"] > 0:
		data += '<tr><td ><b> TOTAL STOCK</b></td>'
		data += '<td align = "center" ><b>%s</b></td>' %(stocks["actual_qty"])
  
	all_whouse = frappe.db.get_value("Warehouse",{"company":company,"is_allocate":1},['name'])
	all_qty = frappe.db.get_value("Bin",{"item_code":item_code,"warehouse":all_whouse},["actual_qty"]) or 0
 
	data += '<tr><td ><b>ALLOCATE QTY</b></td>'
	data += '<td align = "center" ><b>%s</b></td>' %(all_qty)
	
	
	free_qty = stocks["actual_qty"] - all_qty
	data += '<tr><td ><b>FREE QTY</b></td>'
	data += '<td align = "center" ><b>%s</b></td>' %(free_qty)
 
	
	new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' """ % (item_code,company), as_dict=True)[0]
	if not new_po['qty']:
		new_po['qty'] = 0
	if not new_po['d_qty']:
		new_po['d_qty'] = 0
	ppoc_total = new_po['qty'] - new_po['d_qty']

	data += '<tr><td ><b>PURCHASE ORDER QTY</b></td>'
	data += '<td align = "center" ><b>%s</b></td>' %(ppoc_total)
 
	new_so = frappe.db.sql("""select sum(`tabSales Order Item`.qty) as qty,sum(`tabSales Order Item`.delivered_qty) as d_qty from `tabSales Order`
			left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent
			where `tabSales Order Item`.item_code = '%s' and company = '%s' """ % (item_code,company), as_dict=True)[0]
	if not new_so['qty']:
		new_so['qty'] = 0
	if not new_so['d_qty']:
		new_so['d_qty'] = 0
	so_qty = new_so['qty'] - new_so['d_qty']
	
	data += '<tr><td ><b>SALES ORDER QTY</b></td>'
	data += '<td align = "center" ><b>%s</b></td>' %(so_qty)
 
	data += '</table>'
	return data

@frappe.whitelist()
def stock_detail(doc):
	item_details = doc.items
	data = ''
	data += '<h4><center><b>NON AVAILABLE ITEMS</b></center></h4>'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e20026;color:white;"><center><b>ITEM CODE</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e20026;color:white;"><center><b>ITEM NAME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e20026;color:white;"><center><b>ORDERED QTY</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e20026;color:white;"><center><b>NON AVAILABLE QTY</b></center></td></tr>'
	for j in item_details:
		
		
		st = 0
		ware = frappe.db.get_list("Warehouse",{"company":doc.company},['name'])
		for w in ware:
			sto = frappe.db.get_value("Bin",{"item_code":j.item_code,"warehouse":w.name},['actual_qty'])
			if sto and sto>0:
				st+=sto
				
		if st < j.qty:
			a=j.qty-st
			data += '<tr><td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(j.item_code)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(j.item_name)
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(j.qty)
		
			data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' %(a)
		
	data += '</tr>'
	data += '</table>'
	return data

@frappe.whitelist()
def stock_uae(item_details,company):
	item_details = json.loads(item_details)
	frappe.errprint(item_details)
	data = ''
	data += '<h4><center><b>STOCK DETAILS - UAE</b></center></h4>'
	data += '<table class="table table-bordered">'
	data += '<tr>'
	data += '<td colspan=1 style="width:13%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM CODE</b></center></td>'
	data += '<td colspan=1 style="width:33%;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ITEM NAME</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>DESCRIPTION</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>UNIT</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>ALLOCATED QTY</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>FREE QTY</b></center></td>'
	data += '<td colspan=1 style="width:70px;padding:1px;border:1px solid black;font-size:14px;font-size:12px;background-color:#e35310;color:white;"><center><b>PO QTY</b></center></td>'
	for j in item_details:
		stocks = frappe.db.sql("""select sum(`tabBin`.actual_qty) as actual_qty from `tabBin`
							join `tabWarehouse` on `tabWarehouse`.name = `tabBin`.warehouse
							join `tabCompany` on `tabCompany`.name = `tabWarehouse`.company
							where `tabBin`.item_code = '%s' and `tabWarehouse`.company = '%s' """ % (j["item_code"],company), as_dict=True)[0]
		
		if not stocks['actual_qty']:
			stocks['actual_qty'] = 0
			
		all_whouse = frappe.db.get_value("Warehouse",{"company":company,"is_allocate":1},['name'])
		all_qty = frappe.db.get_value("Bin",{"item_code":j["item_code"],"warehouse":all_whouse},["actual_qty"]) or 0
		
		free_qty = stocks["actual_qty"]-all_qty

		new_po = frappe.db.sql("""select sum(`tabPurchase Order Item`.qty) as qty,sum(`tabPurchase Order Item`.received_qty) as d_qty from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus = 1 and `tabPurchase Order`.company = '%s' """ % (j["item_code"],company), as_dict=True)[0]
		if not new_po['qty']:
			new_po['qty'] = 0
		if not new_po['d_qty']:
			new_po['d_qty'] = 0
		ppoc_total = new_po['qty'] - new_po['d_qty']
	 
		data += '<tr>'
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_code"])
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["item_name"])
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["description"])
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (j["uom"])
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (all_qty or 0)
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (free_qty or 0)
		data += '<td style="text-align:center;border: 1px solid black" colspan=1>%s</td>' % (ppoc_total or 0)
		data += '</tr>'
	data += '</table>'
	return data
